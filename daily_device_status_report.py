#!/usr/bin/env python3
"""
DAILY CASING DEVICE & CAMERA STATUS REPORT - AUTOMATED (GitHub Actions)
========================================================================
Runs daily at 6:00 AM Central via GitHub Actions.

Pulls gateway/camera connectivity data from:
1. Motive API /v1/vehicles + /v1/vehicle_locations (vehicle info, locations)
2. Motive Device Status Report CSV (fetched automatically from Gmail via IMAP)

The CSV provides the actual Motive device STATUS field (ground truth for
Powered Off / Camera Powered Off / Inactive 30+ Days classification).
Motive must be configured to email the Device Status Report to the Gmail
account before this script runs (recommended: 5 AM Central daily).

If the email/CSV is unavailable, falls back to API-only mode with a 72-hour
inactivity threshold (catches most issues but can't detect camera-only problems).

Tiered email distribution:
  - Camera Team: daily full report (all yards)
  - Safety: daily full report + escalation summary
  - Dispatch: daily per-yard report (each yard gets only their section)
  - Managers: 7+ day unresolved issues only (per-yard, escalation)

Generates an Excel workbook (.xlsx) + HTML emails matching the BRHAS brand format.
"""

import csv
import email as email_lib
import imaplib
import io
import json
import os
import re
import smtplib
import sys
import zipfile  # for extracting CSV from Motive email ZIP attachments
from collections import OrderedDict
from datetime import datetime, timedelta, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from html import escape as html_escape

import requests

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    from zoneinfo import ZoneInfo
    CENTRAL_TZ = ZoneInfo("America/Chicago")
except Exception:
    CENTRAL_TZ = timezone(timedelta(hours=-6))

# ==============================================================================
# CONFIGURATION
# ==============================================================================

MOTIVE_API_KEY = os.environ.get("MOTIVE_API_KEY")
if not MOTIVE_API_KEY:
    print("ERROR: MOTIVE_API_KEY environment variable is not set.")
    sys.exit(1)

MOTIVE_BASE_URL = "https://api.gomotive.com/v1"

# Thresholds
POWERED_OFF_HOURS = 24       # Hours without activity = powered off
INACTIVE_DAYS = 30           # Days without camera activity = inactive 30+
NEW_THRESHOLD_DAYS = 2       # Disconnected < 2 days = "NEW" on the report
MANAGER_ESCALATION_DAYS = 7  # 7+ days disconnected = escalate to managers
OOS_RECHECK_HOURS = 48       # OOS vehicle with GPS activity within 48h = flag

# ==============================================================================
# RECIPIENT CONFIGURATION (Tiered Workflow)
# ==============================================================================
# Override any tier via environment variables (comma-separated emails).
# Format: DEVICE_REPORT_{TIER}_RECIPIENTS
#
# Tiers:
#   camera_team - Daily full report (all yards, all issues)
#   safety      - Daily full report + escalation summary
#   dispatch    - Daily per-yard report (each yard only sees their issues)
#   managers    - Only 7+ day unresolved issues (per-yard escalation)
# ==============================================================================

# Camera Team: monitors 24/7, gets full daily report (all yards)
CAMERA_TEAM_RECIPIENTS = os.environ.get(
    "DEVICE_REPORT_CAMERA_RECIPIENTS",
    "rdavis@brhas.com, lstromile@brhas.com"
)

# Safety Director: daily full report (all yards) + escalation summary
SAFETY_DIRECTOR_RECIPIENTS = os.environ.get(
    "DEVICE_REPORT_DIRECTOR_RECIPIENTS",
    "krhodes@brhas.com"
)

# Per-yard Safety Reps: daily, their yard(s) only
# Same reps as the speeding report distribution
SAFETY_REP_RECIPIENTS = {
    "Midland":    os.environ.get("DEVICE_REPORT_SAFETY_MIDLAND",    "mhancock@brhas.com, msalazar@brhas.com"),
    "Bryan":      os.environ.get("DEVICE_REPORT_SAFETY_BRYAN",      "jconrad@brhas.com"),
    "Kilgore":    os.environ.get("DEVICE_REPORT_SAFETY_KILGORE",    "jbarnett@brhas.com"),
    "Hobbs":      os.environ.get("DEVICE_REPORT_SAFETY_HOBBS",      "abatts@brhas.com"),
    "Jourdanton": os.environ.get("DEVICE_REPORT_SAFETY_JOURDANTON", "jspeyrer@brhas.com"),
    "Laredo":     os.environ.get("DEVICE_REPORT_SAFETY_LAREDO",     "jspeyrer@brhas.com"),
    "San Angelo": os.environ.get("DEVICE_REPORT_SAFETY_SANANGELO",  "mhancock@brhas.com, msalazar@brhas.com"),
}

# Per-yard Dispatch recipients: daily, their yard only
DISPATCH_RECIPIENTS = {
    "Midland":    os.environ.get("DEVICE_REPORT_DISPATCH_MIDLAND",    "midland.dispatch@brhas.com"),
    "Bryan":      os.environ.get("DEVICE_REPORT_DISPATCH_BRYAN",      "brdispatch@brhas.com"),
    "Kilgore":    os.environ.get("DEVICE_REPORT_DISPATCH_KILGORE",    "tyler.dispatch@brhas.com"),
    "Hobbs":      os.environ.get("DEVICE_REPORT_DISPATCH_HOBBS",      "hobbs.dispatch@brhas.com"),
    "Jourdanton": os.environ.get("DEVICE_REPORT_DISPATCH_JOURDANTON", "pleasantondispatch@brhas.com, laredodispatch@brhas.com"),
    "Laredo":     os.environ.get("DEVICE_REPORT_DISPATCH_LAREDO",     "laredodispatch@brhas.com"),
    "San Angelo": os.environ.get("DEVICE_REPORT_DISPATCH_SANANGELO",  "midland.dispatch@brhas.com"),
}

# Per-yard Manager recipients: ESCALATION ONLY (7+ days unresolved)
# CC'd: mbuffington@brhas.com (ops) on all manager escalation emails
MANAGER_ESCALATION_CC = os.environ.get(
    "DEVICE_REPORT_MANAGER_CC",
    "mbuffington@brhas.com"
)

MANAGER_RECIPIENTS = {
    "Midland":    os.environ.get("DEVICE_REPORT_MANAGER_MIDLAND",    "dustin.fry@brhas.com"),
    "Bryan":      os.environ.get("DEVICE_REPORT_MANAGER_BRYAN",      "eddie.lohse@brhas.com"),
    "Kilgore":    os.environ.get("DEVICE_REPORT_MANAGER_KILGORE",    "fbalderas@brhas.com"),
    "Hobbs":      os.environ.get("DEVICE_REPORT_MANAGER_HOBBS",      "ceaves@brhas.com"),
    "Jourdanton": os.environ.get("DEVICE_REPORT_MANAGER_JOURDANTON", "esflores@brhas.com"),
    "Laredo":     os.environ.get("DEVICE_REPORT_MANAGER_LAREDO",     "cjacobo@brhas.com"),
    "San Angelo": os.environ.get("DEVICE_REPORT_MANAGER_SANANGELO",  "dustin.fry@brhas.com"),
}

# ==============================================================================
# CASING GROUP IDS
# ==============================================================================

CASING_GROUP_IDS = {
    167175: "Midland",
    169090: "Bryan",
    169092: "Kilgore",
    186740: "Hobbs",
    169091: "Jourdanton",
    186739: "Laredo",
    186741: "San Angelo",
    186746: "",  # Parent "Casing" group
}

ALL_CASING_GROUP_IDS = set(CASING_GROUP_IDS.keys())

YARD_ORDER = ["Midland", "Bryan", "Kilgore", "Hobbs", "Jourdanton", "Laredo", "San Angelo"]

# ==============================================================================
# BRHAS THEME COLORS
# ==============================================================================

DARK_BLUE = "1F3864"
MED_BLUE = "2E75B6"
ALT_ROW = "F2F2F2"
RED = "C00000"
ORANGE = "ED7D31"
GOLD = "FFD966"
OOS_BG = "FCE4D6"
WHITE = "FFFFFF"
BLACK = "000000"

# Excel style objects
FILL_DARK_BLUE = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
FILL_MED_BLUE = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
FILL_ALT_ROW = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
FILL_RED = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
FILL_ORANGE = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
FILL_GOLD = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
FILL_OOS = PatternFill(start_color=OOS_BG, end_color=OOS_BG, fill_type="solid")

FONT_WHITE_BOLD = Font(name="Calibri", size=10, bold=True, color=WHITE)
FONT_WHITE_BOLD_14 = Font(name="Calibri", size=14, bold=True, color=WHITE)
FONT_WHITE_10 = Font(name="Calibri", size=10, color=WHITE)
FONT_BLACK_BOLD = Font(name="Calibri", size=10, bold=True, color=BLACK)
FONT_BLACK = Font(name="Calibri", size=10, color=BLACK)
FONT_RED_BOLD = Font(name="Calibri", size=10, bold=True, color=RED)
FONT_DARK_BOLD = Font(name="Calibri", size=10, bold=True, color=BLACK)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ==============================================================================
# MOTIVE API - FETCH CASING VEHICLES
# ==============================================================================

def fetch_casing_vehicles():
    """Fetch all vehicles and filter to Casing division.

    Returns dict keyed by vehicle ID:
        {vehicle_id: {number, make, model, year, yard, driver, eld_device_id,
                      eld_identifier, availability_status, group_ids}}
    """
    headers = {"X-Api-Key": MOTIVE_API_KEY}
    vehicles = {}
    page = 1

    while True:
        try:
            resp = requests.get(
                f"{MOTIVE_BASE_URL}/vehicles",
                headers=headers,
                params={"per_page": 100, "page_no": page},
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
            batch = data.get("vehicles", [])
            if not batch:
                break

            for wrapper in batch:
                v = wrapper.get("vehicle", wrapper)
                vid = v.get("id")
                if not vid:
                    continue

                group_ids = v.get("group_ids", [])
                yard = None
                for gid in group_ids:
                    if gid in CASING_GROUP_IDS:
                        y = CASING_GROUP_IDS[gid]
                        if y:  # Skip empty parent group name
                            yard = y
                            break
                    elif gid in ALL_CASING_GROUP_IDS:
                        yard = ""  # In parent casing group but no specific yard

                if yard is None:
                    continue  # Not a casing vehicle

                # Driver name
                driver = None
                for field in ("current_driver", "permanent_driver"):
                    d = v.get(field)
                    if d and isinstance(d, dict):
                        name = f"{d.get('first_name', '')} {d.get('last_name', '')}".strip()
                        if name:
                            driver = name
                            break

                eld = v.get("eld_device") or {}
                avail = v.get("availability_details") or {}

                vehicles[vid] = {
                    "number": v.get("number", "").strip(),
                    "make": v.get("make", ""),
                    "model": v.get("model", ""),
                    "year": v.get("year", ""),
                    "yard": yard if yard else "Unassigned",
                    "driver": driver or "",
                    "eld_device_id": eld.get("id"),
                    "eld_identifier": eld.get("identifier", ""),
                    "availability": avail.get("availability_status", "unknown"),
                    "group_ids": group_ids,
                }

            pag = data.get("pagination", {})
            if page * 100 >= pag.get("total", 0):
                break
            page += 1

        except Exception as e:
            print(f"  Warning: vehicle fetch page {page} failed: {e}")
            break

    return vehicles


def fetch_vehicle_locations():
    """Fetch last-known locations for all vehicles.

    Returns dict: {vehicle_id: {located_at (datetime), description (str), lat, lon}}
    """
    headers = {"X-Api-Key": MOTIVE_API_KEY}
    locations = {}
    page = 1

    while True:
        try:
            resp = requests.get(
                f"{MOTIVE_BASE_URL}/vehicle_locations",
                headers=headers,
                params={"per_page": 100, "page_no": page},
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
            batch = data.get("vehicles", [])
            if not batch:
                break

            for wrapper in batch:
                v = wrapper.get("vehicle", wrapper)
                vid = v.get("id")
                loc = v.get("current_location") or {}
                located_at_str = loc.get("located_at", "")
                if vid and located_at_str:
                    try:
                        located_at = datetime.fromisoformat(
                            located_at_str.replace("Z", "+00:00")
                        )
                    except Exception:
                        located_at = None

                    locations[vid] = {
                        "located_at": located_at,
                        "description": loc.get("description", ""),
                        "lat": loc.get("lat"),
                        "lon": loc.get("lon"),
                    }

            pag = data.get("pagination", {})
            if page * 100 >= pag.get("total", 0):
                break
            page += 1

        except Exception as e:
            print(f"  Warning: location fetch page {page} failed: {e}")
            break

    return locations


# ==============================================================================
# GMAIL IMAP - FETCH DEVICE STATUS CSV FROM EMAIL
# ==============================================================================

def fetch_csv_from_gmail():
    """Connect to Gmail via IMAP and download the most recent Motive
    Device Status Report and Cameras Status Detail Report attachments.

    Searches for emails matching Motive's subject line patterns:
      - "Your Device Status Report Report for <dates> is ready"
      - "Your Cameras Status Detail Report for <dates> is ready"

    Emails come from notifications@gomotive.com (or any sender for testing).

    Returns tuple: (device_csv: io.StringIO or None, camera_csv: io.StringIO or None)
    """
    gmail_address = os.environ.get("GMAIL_ADDRESS", "")
    gmail_app_password = os.environ.get("GMAIL_APP_PASSWORD", "")

    if not gmail_address or not gmail_app_password:
        print("    GMAIL_ADDRESS or GMAIL_APP_PASSWORD not set -- cannot fetch email.")
        return None, None

    device_csv = None
    camera_csv = None

    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(gmail_address, gmail_app_password)
        mail.select("inbox")

        since_date = (datetime.now() - timedelta(days=3)).strftime("%d-%b-%Y")

        # Search for BOTH report types
        searches = [
            ("Device Status Report", "device"),
            ("Cameras Status Detail", "camera"),
        ]

        for search_term, report_type in searches:
            search_query = f'(SINCE {since_date} SUBJECT "{search_term}")'
            status, msg_ids = mail.search(None, search_query)

            if status != "OK" or not msg_ids[0]:
                print(f"    No '{search_term}' email found in last 3 days.")
                continue

            # Get the most recent matching email
            email_ids = msg_ids[0].split()
            latest_id = email_ids[-1]

            status, msg_data = mail.fetch(latest_id, "(RFC822)")
            if status != "OK":
                print(f"    Failed to fetch {report_type} email.")
                continue

            raw_email = msg_data[0][1]
            msg = email_lib.message_from_bytes(raw_email)

            email_date = msg.get("Date", "unknown")
            email_subject = msg.get("Subject", "unknown")
            email_from = msg.get("From", "unknown")
            print(f"    Found {report_type} email: {email_subject}")
            print(f"      From: {email_from} | Date: {email_date}")

            # Extract CSV from attachments
            csv_content = _extract_csv_from_email(msg, report_type)
            if csv_content:
                if report_type == "device":
                    device_csv = io.StringIO(csv_content)
                else:
                    camera_csv = io.StringIO(csv_content)

        mail.logout()
        return device_csv, camera_csv

    except imaplib.IMAP4.error as e:
        print(f"    IMAP login failed: {e}")
        print("    Make sure IMAP is enabled in Gmail settings and app password is correct.")
        return None, None
    except Exception as e:
        print(f"    Error fetching email: {e}")
        return None, None


def _extract_csv_from_email(msg, report_type):
    """Extract CSV content from an email's attachments (ZIP or direct CSV).

    Returns CSV content as string, or None.
    """
    for part in msg.walk():
        content_type = part.get_content_type()
        filename = part.get_filename()

        if filename is None:
            continue

        payload = part.get_payload(decode=True)
        if payload is None:
            continue

        # Handle ZIP attachments (Motive sends ZIPs)
        if filename.lower().endswith(".zip") or content_type == "application/zip":
            try:
                with zipfile.ZipFile(io.BytesIO(payload)) as zf:
                    csv_names = [
                        n for n in zf.namelist()
                        if n.lower().endswith(".csv")
                    ]
                    if csv_names:
                        csv_bytes = zf.read(csv_names[0])
                        csv_content = csv_bytes.decode("utf-8-sig")
                        print(f"      Extracted from ZIP: {csv_names[0]}")
                        return csv_content
            except Exception as e:
                print(f"      Warning: failed to extract ZIP {filename}: {e}")

        # Handle direct CSV attachments
        elif filename.lower().endswith(".csv"):
            csv_content = payload.decode("utf-8-sig")
            print(f"      Found CSV: {filename}")
            return csv_content

    print(f"      No CSV found in {report_type} email attachments.")
    return None


def parse_device_csv(csv_source):
    """Parse Motive Device Status CSV into per-vehicle device records.

    The CSV has one row per device (Gateway + Camera per vehicle).

    Returns dict: {vehicle_number: {gateway: {...}, camera: {...}}}
    """
    if csv_source is None:
        return {}

    try:
        reader = csv.DictReader(csv_source)

        devices = {}
        for row in reader:
            entity_id = row.get("ENTITY ID", "").strip()
            device_type = row.get("DEVICE", "").strip()
            status = row.get("STATUS", "").strip()
            sub_status = row.get("SUB STATUS", "").strip()
            last_activity = row.get("DEVICE LAST ACTIVITY", "").strip()
            groups = row.get("GROUPS", "").strip()
            power = row.get("POWER", "").strip()

            if not entity_id:
                continue

            if entity_id not in devices:
                devices[entity_id] = {"gateway": None, "camera": None}

            record = {
                "status": status,
                "sub_status": sub_status,
                "last_activity": last_activity,
                "power": power,
                "groups": groups,
            }

            if "gateway" in device_type.lower() or "vehicle gateway" in device_type.lower():
                devices[entity_id]["gateway"] = record
            elif "dashcam" in device_type.lower() or "camera" in device_type.lower():
                devices[entity_id]["camera"] = record

        return devices

    except Exception as e:
        print(f"  Warning: CSV parse failed: {e}")
        return {}


# ==============================================================================
# IFTA MILEAGE (for OOS anomaly detection)
# ==============================================================================

def fetch_oos_mileage(oos_vehicle_numbers, days=7):
    """Fetch recent IFTA trip miles for OOS-active vehicles.

    Returns dict: vehicle_number -> total miles in the last N days.
    """
    if not oos_vehicle_numbers:
        return {}

    headers = {"X-Api-Key": MOTIVE_API_KEY}
    now = datetime.now(timezone.utc).astimezone(CENTRAL_TZ)
    start_date = (now - timedelta(days=days)).strftime("%Y-%m-%d")
    end_date = now.strftime("%Y-%m-%d")

    vehicle_miles = {}
    page = 1
    while True:
        try:
            resp = requests.get(
                f"{MOTIVE_BASE_URL}/ifta/trips",
                headers=headers,
                params={
                    "per_page": 100,
                    "page_no": page,
                    "start_date": start_date,
                    "end_date": end_date,
                },
                timeout=60,
            )
            resp.raise_for_status()
            data = resp.json()
            trips = data.get("ifta_trips", [])
            if not trips:
                break

            for wrapper in trips:
                trip = wrapper.get("ifta_trip", wrapper)
                vehicle = trip.get("vehicle", {})
                vnum = vehicle.get("number", "") if isinstance(vehicle, dict) else str(vehicle)
                if vnum in oos_vehicle_numbers:
                    miles = trip.get("distance", 0) or 0
                    vehicle_miles[vnum] = vehicle_miles.get(vnum, 0) + miles

            pag = data.get("pagination", {})
            total_count = pag.get("total", 0)
            if page * 100 >= total_count:
                break
            page += 1

        except Exception as e:
            print(f"    Warning: IFTA trips page {page} failed: {e}")
            break

    return {vnum: round(miles, 1) for vnum, miles in vehicle_miles.items()}


# ==============================================================================
# ISSUE CLASSIFICATION
# ==============================================================================

def classify_issues(vehicles, locations, csv_devices):
    """Classify each casing vehicle's device status.

    When CSV data is available, uses the actual Motive device STATUS field
    as ground truth (much more accurate than timestamp-based inference).
    When CSV is not available, falls back to API-only mode with a conservative
    72-hour threshold.

    Each issue includes:
      - days_disconnected: integer days since last activity (for display)
      - is_new: True if disconnected < 2 days (just appeared on report)
      - is_escalation: True if disconnected >= 7 days (manager attention)
      - oos_active: True if OOS but showing recent GPS (status update needed)

    Returns list of issue dicts (only vehicles WITH issues).
    """
    now = datetime.now(timezone.utc)
    issues = []
    has_csv = bool(csv_devices)

    # Build a lookup from vehicle number -> CSV data
    csv_lookup = {}
    for entity_id, devs in csv_devices.items():
        clean_id = entity_id.strip()
        csv_lookup[clean_id] = devs

    # Also index CSV by partial match (CSV entity IDs may differ from API)
    csv_by_prefix = {}
    for entity_id in csv_lookup:
        match = re.match(r"^(\d+\s*C?\s*)", entity_id)
        if match:
            prefix = match.group(1).strip()
            csv_by_prefix[prefix] = csv_lookup[entity_id]

    for vid, vinfo in vehicles.items():
        vnum = vinfo["number"]
        yard = vinfo["yard"]
        avail = vinfo["availability"]
        avail_display = "In Service" if avail == "in_service" else "Out Of Service"

        loc = locations.get(vid, {})
        located_at = loc.get("located_at")
        location_desc = loc.get("description", "")

        hours_since = None
        days_since = None
        last_active_str = "Unknown"

        if located_at:
            hours_since = (now - located_at).total_seconds() / 3600
            days_since = hours_since / 24
            last_active_str = located_at.astimezone(CENTRAL_TZ).strftime(
                "%m/%d/%Y %I:%M %p"
            )

        # Try to find CSV data for this vehicle
        csv_data = csv_lookup.get(vnum)
        if not csv_data:
            num_match = re.match(r"^(\d+\s*C?\s*)", vnum)
            if num_match:
                csv_data = csv_by_prefix.get(num_match.group(1).strip())

        gw_csv = csv_data.get("gateway") if csv_data else None
        cam_csv = csv_data.get("camera") if csv_data else None

        def _build_issue(issue_type, devices, action, inactive_days, last_str, last_dt):
            """Helper to build issue dict with NEW/escalation/OOS flags."""
            d = inactive_days if inactive_days is not None else 999
            days_disc = max(0, round(d))

            # OOS re-check: vehicle marked OOS but has recent GPS activity
            oos_active = False
            if avail_display == "Out Of Service" and hours_since is not None:
                if hours_since < OOS_RECHECK_HOURS:
                    oos_active = True

            return {
                "vehicle_number": vnum,
                "yard": yard,
                "availability": avail_display,
                "location": location_desc,
                "last_active": last_str,
                "last_active_dt": last_dt,
                "issue_type": issue_type,
                "devices_affected": devices,
                "recommended_action": action,
                "days_inactive": d,
                "days_disconnected": days_disc,
                "is_new": d < NEW_THRESHOLD_DAYS,
                "is_escalation": d >= MANAGER_ESCALATION_DAYS,
                "oos_active": oos_active,
                "recent_miles": 0,
            }

        # ===== CSV-BASED CLASSIFICATION (ground truth) =====
        if has_csv and (gw_csv or cam_csv):
            gw_status = gw_csv.get("status", "") if gw_csv else ""
            cam_status = cam_csv.get("status", "") if cam_csv else ""
            cam_sub = cam_csv.get("sub_status", "") if cam_csv else ""
            gw_last = gw_csv.get("last_activity", "") if gw_csv else ""
            cam_last = cam_csv.get("last_activity", "") if cam_csv else ""

            display_last = gw_last or cam_last or last_active_str

            # CASE 1: Gateway Powered Off -> both devices down
            if gw_status == "Powered Off":
                devices = "Gateway + Camera"
                action = "Power loss -- gateway down"
                if cam_csv is None:
                    devices = "Gateway Only"
                    action = "Power loss -- no camera record"

                issues.append(_build_issue(
                    "Powered Off", devices, action,
                    days_since, display_last, located_at,
                ))
                continue

            # CASE 2: Gateway OK but Camera Powered Off (not due to gateway)
            if (gw_status == "Normal" and cam_status == "Powered Off"
                    and "gateway powered off" not in cam_sub.lower()):
                issues.append(_build_issue(
                    "Camera Powered Off", "Camera Only",
                    "Camera cable disconnected -- gateway OK",
                    days_since, cam_last or display_last, located_at,
                ))
                continue

            # CASE 3: Both "Normal" but camera inactive 30+ days
            if cam_status == "Normal" and cam_last:
                try:
                    cam_dt = datetime.strptime(cam_last, "%m/%d/%Y %I:%M %p")
                    cam_dt = cam_dt.replace(tzinfo=CENTRAL_TZ)
                    cam_days = (now - cam_dt).total_seconds() / 86400
                    if cam_days >= INACTIVE_DAYS:
                        power = cam_csv.get("power", "") if cam_csv else ""
                        action = "Camera inactive >30 days"
                        if power and "low" in power.lower():
                            action = "Inactive >30 days -- LOW BATTERY"
                        issues.append(_build_issue(
                            "Inactive 30+ Days", "Camera Only", action,
                            cam_days, cam_last, cam_dt,
                        ))
                        continue
                except Exception:
                    pass

            # CASE 4: Both "Normal" but gateway inactive 30+ days
            if gw_status == "Normal" and gw_last:
                try:
                    gw_dt = datetime.strptime(gw_last, "%m/%d/%Y %I:%M %p")
                    gw_dt = gw_dt.replace(tzinfo=CENTRAL_TZ)
                    gw_days = (now - gw_dt).total_seconds() / 86400
                    if gw_days >= INACTIVE_DAYS:
                        power = gw_csv.get("power", "") if gw_csv else ""
                        action = "Inactive >30 days -- LOW BATTERY" if power and "low" in power.lower() else "Inactive >30 days"
                        issues.append(_build_issue(
                            "Inactive 30+ Days", "Gateway + Camera", action,
                            gw_days, gw_last, gw_dt,
                        ))
                        continue
                except Exception:
                    pass

            continue

        # ===== API-ONLY FALLBACK =====
        if has_csv:
            continue

        API_POWERED_OFF_HOURS = 72

        if not located_at:
            issues.append(_build_issue(
                "Powered Off", "Gateway + Camera",
                "Power loss -- gateway down",
                999, "Unknown", None,
            ))
            continue

        if hours_since >= API_POWERED_OFF_HOURS:
            issues.append(_build_issue(
                "Powered Off", "Gateway + Camera",
                "Power loss -- gateway down",
                days_since, last_active_str, located_at,
            ))

    # Sort: In Service first, then by issue severity, then by days inactive
    def sort_key(issue):
        avail_order = 0 if issue["availability"] == "In Service" else 1
        type_order = {"Powered Off": 0, "Camera Powered Off": 1, "Inactive 30+ Days": 2}
        return (avail_order, type_order.get(issue["issue_type"], 9), -issue["days_inactive"])

    return sorted(issues, key=sort_key)


def group_issues_by_yard(issues):
    """Group issues by yard, following YARD_ORDER."""
    raw = {}
    for issue in issues:
        yard = issue["yard"]
        raw.setdefault(yard, []).append(issue)

    grouped = OrderedDict()
    for yard in YARD_ORDER:
        if yard in raw:
            grouped[yard] = raw[yard]

    for yard in sorted(raw.keys()):
        if yard not in grouped:
            grouped[yard] = raw[yard]

    return grouped


# ==============================================================================
# EXCEL REPORT GENERATION
# ==============================================================================

FILL_GREEN = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
FONT_NEW_TAG = Font(name="Calibri", size=9, bold=True, color="00B050")


def _apply_header_row(ws, row, values, fill, font, max_col=9):
    """Write a merged header row across all columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=values)
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for c in range(2, max_col + 1):
        ws.cell(row=row, column=c).fill = fill


def _apply_table_headers(ws, row, headers, fill, font):
    """Write table header row with individual cells."""
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _issue_type_fill(issue_type):
    """Return the fill color for an issue type cell."""
    if issue_type == "Powered Off":
        return FILL_RED
    elif issue_type == "Camera Powered Off":
        return FILL_ORANGE
    elif issue_type == "Inactive 30+ Days":
        return FILL_GOLD
    return None


def _issue_type_font(issue_type):
    """Return the font for an issue type cell."""
    if issue_type == "Powered Off":
        return Font(name="Calibri", size=10, bold=True, color=WHITE)
    elif issue_type == "Camera Powered Off":
        return Font(name="Calibri", size=10, bold=True, color=WHITE)
    elif issue_type == "Inactive 30+ Days":
        return Font(name="Calibri", size=10, bold=True, color=BLACK)
    return FONT_BLACK


def create_excel_report(issues, grouped, report_date):
    """Build the full Excel workbook with Days Disconnected and NEW/escalation flags."""
    wb = Workbook()

    # ---- SUMMARY SHEET ----
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = DARK_BLUE

    col_widths = [22, 14, 14, 14, 14, 14, 14, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    date_str = report_date.strftime("%B %d, %Y")

    _apply_header_row(
        ws, 1,
        f"CAMERA & GATEWAY DISCONNECTION REPORT  --  As of {date_str}",
        FILL_DARK_BLUE, FONT_WHITE_BOLD_14,
    )
    ws.row_dimensions[1].height = 30

    _apply_header_row(
        ws, 2,
        "Devices that were previously active but are now Powered Off, Camera Powered Off, or Inactive 30+ Days  |  Organized by Casing Group",
        FILL_MED_BLUE, FONT_WHITE_10,
    )
    ws.row_dimensions[2].height = 22

    summary_headers = [
        "Casing Group", "Total Vehicles\nAffected", "Powered Off\n(Full)",
        "Camera Only\nIssues", "Inactive 30+\nDays", "Out of Service\nVehicles",
        "In Service\nVehicles", "NEW\nIssues", "7+ Day\nEscalation",
    ]
    _apply_table_headers(ws, 4, summary_headers, FILL_MED_BLUE, FONT_WHITE_BOLD)
    ws.row_dimensions[4].height = 30

    row = 5
    for yard in YARD_ORDER:
        yard_issues = grouped.get(yard, [])
        if not yard_issues:
            continue

        total = len(yard_issues)
        powered_off = len([i for i in yard_issues if i["issue_type"] == "Powered Off"])
        camera_only = len([i for i in yard_issues if i["issue_type"] in ("Camera Powered Off", "Inactive 30+ Days") and i["devices_affected"] == "Camera Only"])
        inactive_30 = len([i for i in yard_issues if i["issue_type"] == "Inactive 30+ Days"])
        oos = len([i for i in yard_issues if i["availability"] == "Out Of Service"])
        in_service = len([i for i in yard_issues if i["availability"] == "In Service"])
        new_count = len([i for i in yard_issues if i["is_new"]])
        escalation_count = len([i for i in yard_issues if i["is_escalation"]])

        values = [f"{yard} Casing", total, powered_off, camera_only, inactive_30, oos, in_service, new_count, escalation_count]
        is_alt = (row - 5) % 2 == 1

        for i, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.font = FONT_BLACK_BOLD if i == 1 else FONT_BLACK
            cell.alignment = ALIGN_CENTER if i > 1 else ALIGN_LEFT
            cell.border = THIN_BORDER
            if is_alt:
                cell.fill = FILL_ALT_ROW
            # Highlight escalation count in red if > 0
            if i == 9 and val > 0:
                cell.font = FONT_RED_BOLD

        row += 1

    # TOTAL row
    total_row = row
    total_issues = len(issues)
    total_powered = len([i for i in issues if i["issue_type"] == "Powered Off"])
    total_cam_only = len([i for i in issues if i["issue_type"] in ("Camera Powered Off", "Inactive 30+ Days") and i["devices_affected"] == "Camera Only"])
    total_inactive = len([i for i in issues if i["issue_type"] == "Inactive 30+ Days"])
    total_oos = len([i for i in issues if i["availability"] == "Out Of Service"])
    total_in_svc = len([i for i in issues if i["availability"] == "In Service"])
    total_new = len([i for i in issues if i["is_new"]])
    total_escalation = len([i for i in issues if i["is_escalation"]])

    total_values = ["TOTAL", total_issues, total_powered, total_cam_only, total_inactive, total_oos, total_in_svc, total_new, total_escalation]
    for i, val in enumerate(total_values, 1):
        cell = ws.cell(row=total_row, column=i, value=val)
        cell.fill = FILL_DARK_BLUE
        cell.font = FONT_WHITE_BOLD
        cell.alignment = ALIGN_CENTER if i > 1 else ALIGN_LEFT
        cell.border = THIN_BORDER

    # ---- PER-YARD SHEETS ----
    detail_headers = [
        "Vehicle ID", "Availability", "Last Known Location",
        "Last Active", "Days\nDisconnected", "Status",
        "Issue Type", "Device(s) Affected", "Recommended Action",
    ]
    detail_widths = [30, 16, 35, 22, 12, 10, 18, 20, 40]

    for yard in YARD_ORDER:
        yard_issues = grouped.get(yard, [])
        if not yard_issues:
            continue

        ws = wb.create_sheet(title=yard)
        ws.sheet_properties.tabColor = DARK_BLUE

        for i, w in enumerate(detail_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        _apply_header_row(
            ws, 1,
            f"{yard} Casing  --  Disconnected / Inactive Devices",
            FILL_DARK_BLUE, FONT_WHITE_BOLD_14,
        )
        ws.row_dimensions[1].height = 30

        powered_off = len([i for i in yard_issues if i["issue_type"] == "Powered Off"])
        camera_disc = len([i for i in yard_issues if i["issue_type"] == "Camera Powered Off"])
        inactive_30 = len([i for i in yard_issues if i["issue_type"] == "Inactive 30+ Days"])
        new_count = len([i for i in yard_issues if i["is_new"]])
        esc_count = len([i for i in yard_issues if i["is_escalation"]])
        stats_text = (
            f"Total: {len(yard_issues)}  |  "
            f"Powered Off: {powered_off}  |  "
            f"Camera Disconnected: {camera_disc}  |  "
            f"Inactive 30+: {inactive_30}  |  "
            f"NEW: {new_count}  |  "
            f"7+ Day Escalation: {esc_count}"
        )
        _apply_header_row(ws, 2, stats_text, FILL_MED_BLUE, Font(name="Calibri", size=10, bold=True, color=WHITE))
        ws.row_dimensions[2].height = 22

        _apply_table_headers(ws, 4, detail_headers, FILL_MED_BLUE, FONT_WHITE_BOLD)
        ws.row_dimensions[4].height = 22

        data_row = 5
        for idx, issue in enumerate(yard_issues):
            is_alt = idx % 2 == 1
            row_fill = FILL_ALT_ROW if is_alt else None

            # Status label
            if issue["is_new"]:
                status_label = "NEW"
            elif issue["is_escalation"]:
                status_label = "ESCALATION"
            else:
                status_label = ""

            # OOS with recent activity override
            avail_text = issue["availability"]
            if issue["oos_active"]:
                avail_text = "OOS - Active"

            values = [
                issue["vehicle_number"],
                avail_text,
                issue["location"],
                issue["last_active"],
                issue["days_disconnected"],
                status_label,
                issue["issue_type"],
                issue["devices_affected"],
                issue["recommended_action"],
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.alignment = ALIGN_LEFT

                if col == 1:  # Vehicle ID
                    cell.font = FONT_BLACK_BOLD
                    if row_fill:
                        cell.fill = row_fill
                elif col == 2:  # Availability
                    if issue["oos_active"]:
                        cell.fill = FILL_ORANGE
                        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
                    elif val == "Out Of Service":
                        cell.fill = FILL_OOS
                        cell.font = FONT_RED_BOLD
                    else:
                        cell.font = FONT_BLACK
                        if row_fill:
                            cell.fill = row_fill
                elif col == 5:  # Days Disconnected
                    cell.alignment = ALIGN_CENTER
                    cell.font = FONT_RED_BOLD if issue["is_escalation"] else FONT_BLACK_BOLD
                    if row_fill:
                        cell.fill = row_fill
                elif col == 6:  # Status (NEW / ESCALATION)
                    cell.alignment = ALIGN_CENTER
                    if val == "NEW":
                        cell.font = Font(name="Calibri", size=10, bold=True, color="00B050")
                    elif val == "ESCALATION":
                        cell.font = FONT_RED_BOLD
                    else:
                        cell.font = FONT_BLACK
                    if row_fill:
                        cell.fill = row_fill
                elif col == 7:  # Issue Type -- color-coded
                    cell.fill = _issue_type_fill(val) or (row_fill if row_fill else PatternFill())
                    cell.font = _issue_type_font(val)
                    cell.alignment = ALIGN_CENTER
                else:
                    cell.font = FONT_BLACK
                    if row_fill:
                        cell.fill = row_fill

            data_row += 1

    # ---- ESCALATION SHEET (7+ day issues only) ----
    escalation_issues = [i for i in issues if i["is_escalation"]]
    if escalation_issues:
        ws = wb.create_sheet(title="Escalation (7+ Days)")
        ws.sheet_properties.tabColor = RED

        esc_headers = [
            "Vehicle ID", "Yard", "Availability", "Days\nDisconnected",
            "Issue Type", "Device(s) Affected", "Last Active", "Recommended Action",
        ]
        esc_widths = [30, 16, 16, 12, 18, 20, 22, 40]
        for i, w in enumerate(esc_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        _apply_header_row(
            ws, 1,
            f"ESCALATION -- Devices Disconnected 7+ Days  (as of {date_str})",
            PatternFill(start_color=RED, end_color=RED, fill_type="solid"),
            FONT_WHITE_BOLD_14, max_col=8,
        )
        ws.row_dimensions[1].height = 30

        _apply_header_row(
            ws, 2,
            f"{len(escalation_issues)} vehicle(s) have been disconnected for 7 or more days and require management attention",
            FILL_MED_BLUE, Font(name="Calibri", size=10, bold=True, color=WHITE), max_col=8,
        )
        ws.row_dimensions[2].height = 22

        _apply_table_headers(ws, 4, esc_headers, FILL_MED_BLUE, FONT_WHITE_BOLD)
        ws.row_dimensions[4].height = 22

        data_row = 5
        for idx, issue in enumerate(escalation_issues):
            is_alt = idx % 2 == 1
            row_fill = FILL_ALT_ROW if is_alt else None

            values = [
                issue["vehicle_number"], issue["yard"], issue["availability"],
                issue["days_disconnected"], issue["issue_type"],
                issue["devices_affected"], issue["last_active"],
                issue["recommended_action"],
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.alignment = ALIGN_LEFT
                cell.font = FONT_BLACK
                if col == 1:
                    cell.font = FONT_BLACK_BOLD
                elif col == 4:
                    cell.font = FONT_RED_BOLD
                    cell.alignment = ALIGN_CENTER
                elif col == 5:
                    cell.fill = _issue_type_fill(val) or PatternFill()
                    cell.font = _issue_type_font(val)
                    cell.alignment = ALIGN_CENTER
                if row_fill and col not in (5,):
                    cell.fill = row_fill
            data_row += 1

    # ---- LEGEND SHEET ----
    ws = wb.create_sheet(title="Legend")
    ws.sheet_properties.tabColor = DARK_BLUE

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80

    _apply_header_row(ws, 1, "Color Legend & Status Guide", FILL_DARK_BLUE, FONT_WHITE_BOLD_14, max_col=3)

    for i, h in enumerate(["Label", "Color / Tag", "Description"], 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.fill = FILL_MED_BLUE
        cell.font = FONT_WHITE_BOLD
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    legend_data = [
        ("Powered Off", FILL_RED, Font(name="Calibri", size=10, bold=True, color=WHITE),
         "Gateway AND camera both lost power. Check cable connections, battery kill switch, and ignition power source."),
        ("Camera Powered Off", FILL_ORANGE, Font(name="Calibri", size=10, bold=True, color=WHITE),
         "Gateway is OK -- camera cable is disconnected. Reconnect the dashcam cable to the Vehicle Gateway."),
        ("Inactive 30+ Days", FILL_GOLD, Font(name="Calibri", size=10, bold=True, color=BLACK),
         "Device shows Normal status but has not reported in 30+ days. Check vehicle status and troubleshoot device."),
        ("NEW", PatternFill(), Font(name="Calibri", size=10, bold=True, color="00B050"),
         "Vehicle just appeared on the disconnected list (< 2 days). Recently lost connectivity."),
        ("ESCALATION", PatternFill(), Font(name="Calibri", size=10, bold=True, color=RED),
         "Vehicle has been disconnected for 7+ days without resolution. Requires management attention."),
        ("OOS - Active", FILL_ORANGE, Font(name="Calibri", size=10, bold=True, color=WHITE),
         "Vehicle is marked Out Of Service in Motive but shows recent GPS activity. Status needs to be updated in Motive."),
    ]

    for idx, (label, fill, font, desc) in enumerate(legend_data):
        row = 4 + idx
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = FONT_BLACK_BOLD
        cell_a.border = THIN_BORDER

        cell_b = ws.cell(row=row, column=2, value=label)
        cell_b.fill = fill
        cell_b.font = font
        cell_b.alignment = ALIGN_CENTER
        cell_b.border = THIN_BORDER

        cell_c = ws.cell(row=row, column=3, value=desc)
        cell_c.font = FONT_BLACK
        cell_c.alignment = ALIGN_LEFT
        cell_c.border = THIN_BORDER

    return wb


# ==============================================================================
# HTML EMAIL
# ==============================================================================

def _h(text):
    """HTML-escape text safely."""
    return html_escape(str(text)) if text else ""


def _issue_html_color(issue_type):
    """Return (bg_color, text_color) for issue type."""
    if issue_type == "Powered Off":
        return f"#{RED}", f"#{WHITE}"
    elif issue_type == "Camera Powered Off":
        return f"#{ORANGE}", f"#{WHITE}"
    elif issue_type == "Inactive 30+ Days":
        return f"#{GOLD}", f"#{BLACK}"
    return "#EEEEEE", f"#{BLACK}"


def _build_issue_rows_html(issue_list):
    """Build HTML table rows for a list of issues."""
    rows = []
    for idx, issue in enumerate(issue_list):
        bg = f"#{ALT_ROW}" if idx % 2 == 1 else "#ffffff"
        issue_bg, issue_fg = _issue_html_color(issue["issue_type"])

        avail_text = issue["availability"]
        avail_style = ""
        if issue["oos_active"]:
            avail_text = "OOS - Active"
            avail_style = f"background:#{ORANGE};color:#fff;font-weight:bold;"
        elif issue["availability"] == "Out Of Service":
            avail_style = f"background:#{OOS_BG};color:#{RED};font-weight:bold;"

        if issue["is_new"]:
            status_html = '<span style="color:#00B050;font-weight:bold;">NEW</span>'
        elif issue["is_escalation"]:
            status_html = f'<span style="color:#{RED};font-weight:bold;">ESCALATION</span>'
        else:
            status_html = ""

        days_val = issue["days_disconnected"]
        days_style = f"color:#{RED};font-weight:bold;" if issue["is_escalation"] else "font-weight:bold;"

        rows.append(f"""    <tr style="background:{bg};">
      <td style="padding:5px 8px;border:1px solid #ddd;font-weight:bold;">{_h(issue["vehicle_number"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;{avail_style}">{_h(avail_text)}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;text-align:center;{days_style}">{days_val}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;text-align:center;">{status_html}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(issue["location"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(issue["last_active"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;background:{issue_bg};color:{issue_fg};font-weight:bold;text-align:center;">{_h(issue["issue_type"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(issue["devices_affected"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(issue["recommended_action"])}</td>
    </tr>""")
    return "\n".join(rows)


def _build_issue_table_header_html(header_bg):
    """Build the column header row for an issue table."""
    return f"""  <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;font-size:12px;">
    <tr style="background:{header_bg};">
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Vehicle ID</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Availability</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:center;">Days</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:center;">Status</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Last Location</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Last Active</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:center;">Issue Type</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Device(s)</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Action</th>
    </tr>"""


def _build_issue_table_html(yard_issues, yard_name):
    """Build an HTML table of issues for a yard section (reused across modes).

    Splits in-service and out-of-service into separate sections.
    In-service issues appear first; OOS vehicles appear in a muted
    'For Awareness' section below.
    """
    in_svc = [i for i in yard_issues if i["availability"] == "In Service"]
    oos = [i for i in yard_issues if i["availability"] != "In Service"]

    parts = []
    powered = len([i for i in yard_issues if i["issue_type"] == "Powered Off"])
    cam_disc = len([i for i in yard_issues if i["issue_type"] == "Camera Powered Off"])
    inact = len([i for i in yard_issues if i["issue_type"] == "Inactive 30+ Days"])
    y_new = len([i for i in yard_issues if i["is_new"]])

    parts.append(f"""
<tr><td style="padding:0 40px;"><hr style="border:none;border-top:3px solid #{DARK_BLUE};margin:15px 0 0 0;"></td></tr>
<tr><td style="padding:15px 40px;">
  <h2 style="color:#{DARK_BLUE};margin:0;font-size:18px;">{_h(yard_name.upper())} CASING</h2>
  <div style="background:#{MED_BLUE};color:#fff;padding:8px 15px;margin:8px 0;font-size:12px;font-weight:bold;">
    Total: {len(yard_issues)} | In Service: {len(in_svc)} | Out of Service: {len(oos)} | Powered Off: {powered} | Camera Disconnected: {cam_disc} | Inactive 30+: {inact} | NEW: {y_new}
  </div>""")

    # In-service issues (primary section)
    if in_svc:
        parts.append(_build_issue_table_header_html(f"#{MED_BLUE}"))
        parts.append(_build_issue_rows_html(in_svc))
        parts.append("  </table>")
    else:
        parts.append(f'  <div style="padding:10px 0;color:#666;font-style:italic;">No in-service vehicles with device issues.</div>')

    # OOS issues (awareness section)
    if oos:
        parts.append(f"""
  <div style="background:#999999;color:#fff;padding:6px 15px;margin:15px 0 0 0;font-size:11px;font-weight:bold;">
    OUT OF SERVICE -- FOR AWARENESS ({len(oos)} vehicle{"s" if len(oos) != 1 else ""})
  </div>""")
        parts.append(_build_issue_table_header_html("#999999"))
        parts.append(_build_issue_rows_html(oos))
        parts.append("  </table>")

    parts.append("</td></tr>")
    return "\n".join(parts)


def _build_html_header(header_bg, header_title, date_str, gen_str):
    """Build the common HTML email header/shell."""
    return f"""<html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f4f4f4;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f4;">
<tr><td align="center">
<table width="750" cellpadding="0" cellspacing="0" style="background:#ffffff;border:1px solid #ddd;margin:20px auto;font-family:Calibri,Arial,sans-serif;font-size:14px;color:#333;">

<tr><td style="background:#{header_bg};padding:25px 40px;text-align:center;">
  <div style="font-size:12px;font-weight:bold;color:#ffffff;letter-spacing:1px;">BRHAS CASING DIVISION</div>
  <div style="font-size:24px;font-weight:bold;color:#ffffff;margin:8px 0;">{header_title}</div>
  <div style="font-size:11px;color:#a0b4d0;margin-top:6px;">Report Date: {date_str} | Generated: {gen_str}</div>
</td></tr>"""


def _build_html_footer(mode, gen_str):
    """Build the common HTML email footer."""
    tier_label = {"full": "Daily Report", "yard": "Yard Report",
                  "escalation": "Escalation Alert", "director": "Director Summary"}
    return f"""
<tr><td style="background:#{DARK_BLUE};padding:15px 40px;text-align:center;">
  <div style="color:#ffffff;font-size:11px;">Butch's Rat Hole & Anchor Service Inc. | Casing Division | HSE Department</div>
  <div style="color:#a0b4d0;font-size:10px;margin-top:4px;">Automated {tier_label.get(mode, "Daily Report")} | Generated {gen_str}</div>
</td></tr>

</table>
</td></tr></table>
</body></html>"""


def _build_oos_anomaly_html(oos_active_list):
    """Build the OOS anomalies section with mileage data."""
    parts = []
    # Sort by miles descending -- trucks with real miles are the priority
    sorted_oos = sorted(oos_active_list, key=lambda i: i.get("recent_miles", 0), reverse=True)

    parts.append(f"""
<tr><td style="padding:0 40px;"><hr style="border:none;border-top:3px solid #{ORANGE};margin:15px 0 0 0;"></td></tr>
<tr><td style="padding:15px 40px;">
  <h2 style="color:#{ORANGE};margin:0 0 10px 0;font-size:16px;">OOS ANOMALIES -- {len(sorted_oos)} Vehicle{"s" if len(sorted_oos) != 1 else ""}</h2>
  <div style="font-size:12px;color:#666;margin-bottom:8px;">Marked Out Of Service in Motive but showing recent GPS activity. Mileage shown is from IFTA trip data (last 7 days). High mileage = truck is running jobs and needs status updated.</div>
  <table cellpadding="0" cellspacing="0" style="border-collapse:collapse;font-size:12px;">
    <tr style="background:#{MED_BLUE};">
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;">Vehicle ID</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;">Yard</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;">Last GPS Activity</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;">Miles (7 days)</th>
    </tr>""")

    for i in sorted_oos:
        miles = i.get("recent_miles", 0)
        if miles > 50:
            miles_style = f"color:#{RED};font-weight:bold;"
            miles_text = f"{miles}"
        elif miles > 0:
            miles_style = f"color:#{ORANGE};font-weight:bold;"
            miles_text = f"{miles}"
        else:
            miles_style = "color:#999;"
            miles_text = "0"

        parts.append(f"""    <tr>
      <td style="padding:5px 8px;border:1px solid #ddd;font-weight:bold;">{_h(i["vehicle_number"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(i["yard"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(i["last_active"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;text-align:center;{miles_style}">{miles_text}</td>
    </tr>""")

    parts.append("  </table>\n</td></tr>")
    return "\n".join(parts)


def _build_html_legend():
    """Build the color legend section."""
    return f"""
<tr><td style="padding:0 40px;"><hr style="border:none;border-top:3px solid #{DARK_BLUE};margin:15px 0 0 0;"></td></tr>
<tr><td style="padding:15px 40px;">
  <h2 style="color:#{DARK_BLUE};margin:0 0 10px 0;font-size:16px;">COLOR LEGEND</h2>
  <table cellpadding="0" cellspacing="0" style="border-collapse:collapse;font-size:12px;">
    <tr>
      <td style="background:#{RED};color:#fff;font-weight:bold;padding:6px 12px;border:1px solid #ddd;">Powered Off</td>
      <td style="padding:6px 12px;border:1px solid #ddd;">Gateway AND camera lost power. Check cable connections, battery kill switch, ignition source.</td>
    </tr>
    <tr>
      <td style="background:#{ORANGE};color:#fff;font-weight:bold;padding:6px 12px;border:1px solid #ddd;">Camera Powered Off</td>
      <td style="padding:6px 12px;border:1px solid #ddd;">Gateway OK -- camera cable disconnected. Reconnect dashcam cable to Vehicle Gateway.</td>
    </tr>
    <tr>
      <td style="background:#{GOLD};color:#000;font-weight:bold;padding:6px 12px;border:1px solid #ddd;">Inactive 30+ Days</td>
      <td style="padding:6px 12px;border:1px solid #ddd;">Device shows Normal but no activity in 30+ days. Check vehicle status and troubleshoot.</td>
    </tr>
    <tr>
      <td style="color:#00B050;font-weight:bold;padding:6px 12px;border:1px solid #ddd;">NEW</td>
      <td style="padding:6px 12px;border:1px solid #ddd;">Vehicle just appeared on disconnected list (within 2 days).</td>
    </tr>
    <tr>
      <td style="color:#{RED};font-weight:bold;padding:6px 12px;border:1px solid #ddd;">ESCALATION</td>
      <td style="padding:6px 12px;border:1px solid #ddd;">Disconnected 7+ days. Requires management attention.</td>
    </tr>
  </table>
</td></tr>"""


def build_html_email(issues, grouped, report_date, csv_available,
                     mode="full", yard_filter=None):
    """Build HTML email body.

    Args:
        mode: "full"       = all yards, all issues (camera team)
              "director"   = 30-second summary: numbers + NEW + 7+ day escalations + OOS flags
              "yard"       = single yard only (dispatch, safety rep)
              "escalation" = 7+ day issues only (manager escalation)
        yard_filter: yard name to filter to (for mode="yard" or "escalation")
    """
    now_central = datetime.now(timezone.utc).astimezone(CENTRAL_TZ)
    date_str = report_date.strftime("%A, %B %d, %Y")
    gen_str = now_central.strftime("%B %d, %Y at %I:%M %p CT")

    # =====================================================================
    # DIRECTOR MODE -- 30-second scan email
    # Shows: dashboard numbers, NEW vehicles, 7+ day escalations, OOS flags
    # Skips: the 3-6 day "in progress" middle ground
    # Full Excel attached for drill-down
    # =====================================================================
    if mode == "director":
        return _build_director_email(issues, grouped, report_date, csv_available,
                                     date_str, gen_str)

    # Filter issues based on mode
    if mode == "yard" and yard_filter:
        display_issues = [i for i in issues if i["yard"] == yard_filter]
        display_grouped = OrderedDict()
        if yard_filter in grouped:
            display_grouped[yard_filter] = grouped[yard_filter]
    elif mode == "escalation":
        if yard_filter:
            display_issues = [i for i in issues if i["is_escalation"] and i["yard"] == yard_filter and i["availability"] == "In Service"]
        else:
            display_issues = [i for i in issues if i["is_escalation"] and i["availability"] == "In Service"]
        display_grouped = OrderedDict()
        for yard in YARD_ORDER:
            esc = [i for i in grouped.get(yard, []) if i["is_escalation"] and i["availability"] == "In Service"]
            if yard_filter and yard != yard_filter:
                continue
            if esc:
                display_grouped[yard] = esc
    else:
        display_issues = issues
        display_grouped = grouped

    total = len(display_issues)
    powered_off = len([i for i in display_issues if i["issue_type"] == "Powered Off"])
    camera_only = len([i for i in display_issues if i["devices_affected"] == "Camera Only"])
    inactive_30 = len([i for i in display_issues if i["issue_type"] == "Inactive 30+ Days"])
    in_service_issues = len([i for i in display_issues if i["availability"] == "In Service"])
    new_count = len([i for i in display_issues if i["is_new"]])
    esc_count = len([i for i in display_issues if i["is_escalation"]])

    parts = []

    # Header
    if mode == "escalation":
        header_title = "DEVICE ESCALATION REPORT -- 7+ DAYS UNRESOLVED"
        header_bg = RED
    elif mode == "yard" and yard_filter:
        header_title = f"{yard_filter.upper()} CASING -- DEVICE STATUS"
        header_bg = DARK_BLUE
    else:
        header_title = "CAMERA & GATEWAY DISCONNECTION REPORT"
        header_bg = DARK_BLUE

    parts.append(_build_html_header(header_bg, header_title, date_str, gen_str))

    # Executive Summary
    parts.append(f"""
<tr><td style="padding:20px 40px;">
  <h2 style="color:#{DARK_BLUE};margin:0 0 15px 0;font-size:18px;border-bottom:2px solid #{DARK_BLUE};padding-bottom:5px;">{"ESCALATION SUMMARY" if mode == "escalation" else "EXECUTIVE SUMMARY"}</h2>
  <div style="background:#f8f8f8;border-left:4px solid #{DARK_BLUE};padding:15px;margin:10px 0;">
    <div style="font-size:16px;font-weight:bold;margin-bottom:8px;">{total} Vehicle{"s" if total != 1 else ""} {"Requiring Management Attention" if mode == "escalation" else "with Device Issues"}</div>
    <div style="margin:4px 0;"><span style="display:inline-block;width:12px;height:12px;background:#{RED};margin-right:6px;vertical-align:middle;"></span><b>Powered Off:</b> {powered_off}</div>
    <div style="margin:4px 0;"><span style="display:inline-block;width:12px;height:12px;background:#{ORANGE};margin-right:6px;vertical-align:middle;"></span><b>Camera Only Issues:</b> {camera_only}</div>
    <div style="margin:4px 0;"><span style="display:inline-block;width:12px;height:12px;background:#{GOLD};margin-right:6px;vertical-align:middle;"></span><b>Inactive 30+ Days:</b> {inactive_30}</div>""")

    if mode != "escalation":
        parts.append(f"""    <div style="margin:4px 0;"><span style="display:inline-block;width:12px;height:12px;background:#00B050;margin-right:6px;vertical-align:middle;"></span><b>NEW (last 2 days):</b> {new_count}</div>""")

    if esc_count > 0 and mode != "escalation":
        parts.append(f"""    <div style="margin:8px 0 0 0;font-weight:bold;color:#{RED};">{esc_count} vehicle{"s" if esc_count != 1 else ""} disconnected 7+ days -- ESCALATION</div>""")

    parts.append(f"""    <div style="margin:4px 0;font-weight:bold;color:#{RED};">{in_service_issues} In-Service vehicle{"s" if in_service_issues != 1 else ""} need attention</div>
  </div>""")

    if not csv_available and mode == "full":
        parts.append(f"""
  <div style="background:#FFF8E1;border-left:4px solid #{GOLD};padding:10px 15px;margin:10px 0;font-size:12px;">
    <b>Note:</b> Camera-only disconnections require a Motive Device Status CSV for detection.
    This report covers all gateway-level disconnections which account for the majority of issues.
  </div>""")

    parts.append("</td></tr>")

    # Per-yard sections
    for yard, yard_issues in display_grouped.items():
        parts.append(_build_issue_table_html(yard_issues, yard))

    # OOS-Active vehicles callout (full mode only)
    oos_active_list = [i for i in display_issues if i["oos_active"]]
    if oos_active_list and mode == "full":
        parts.append(_build_oos_anomaly_html(oos_active_list))

    # Legend (skip for escalation-only emails)
    if mode != "escalation":
        parts.append(_build_html_legend())

    parts.append(_build_html_footer(mode, gen_str))
    return "\n".join(parts)


def _build_director_email(issues, grouped, report_date, csv_available,
                          date_str, gen_str):
    """Build the Director Summary email -- 30-second scan.

    Structure:
      1. Dashboard: total issues, by-yard breakdown (one-line per yard)
      2. WHAT'S NEW TODAY: only vehicles disconnected < 2 days (compact table)
      3. ESCALATION (7+ DAYS): vehicles nobody has fixed (compact table)
      4. OOS ANOMALIES: trucks marked OOS but showing GPS activity
      5. Footer note: "Full Excel attached for drill-down"
    """
    parts = []
    total = len(issues)
    new_issues = [i for i in issues if i["is_new"]]
    esc_issues = [i for i in issues if i["is_escalation"]]
    oos_active_list = [i for i in issues if i["oos_active"]]
    in_svc = len([i for i in issues if i["availability"] == "In Service"])

    # Header
    parts.append(_build_html_header(DARK_BLUE, "DIRECTOR DAILY SUMMARY", date_str, gen_str))

    # ---- DASHBOARD ----
    parts.append(f"""
<tr><td style="padding:20px 40px;">
  <h2 style="color:#{DARK_BLUE};margin:0 0 12px 0;font-size:18px;border-bottom:2px solid #{DARK_BLUE};padding-bottom:5px;">DASHBOARD</h2>
  <table cellpadding="0" cellspacing="0" style="border-collapse:collapse;width:100%;margin-bottom:15px;">
    <tr>
      <td style="background:#{DARK_BLUE};color:#fff;padding:18px;text-align:center;width:25%;border:2px solid #fff;">
        <div style="font-size:28px;font-weight:bold;">{total}</div>
        <div style="font-size:11px;margin-top:4px;">TOTAL ISSUES</div>
      </td>
      <td style="background:#00B050;color:#fff;padding:18px;text-align:center;width:25%;border:2px solid #fff;">
        <div style="font-size:28px;font-weight:bold;">{len(new_issues)}</div>
        <div style="font-size:11px;margin-top:4px;">NEW TODAY</div>
      </td>
      <td style="background:#{RED};color:#fff;padding:18px;text-align:center;width:25%;border:2px solid #fff;">
        <div style="font-size:28px;font-weight:bold;">{len(esc_issues)}</div>
        <div style="font-size:11px;margin-top:4px;">ESCALATION (7+ DAYS)</div>
      </td>
      <td style="background:#{ORANGE};color:#fff;padding:18px;text-align:center;width:25%;border:2px solid #fff;">
        <div style="font-size:28px;font-weight:bold;">{len(oos_active_list)}</div>
        <div style="font-size:11px;margin-top:4px;">OOS ANOMALIES</div>
      </td>
    </tr>
  </table>""")

    # Yard breakdown (one line each)
    parts.append(f"""  <table cellpadding="0" cellspacing="0" style="border-collapse:collapse;width:100%;font-size:12px;">
    <tr style="background:#{MED_BLUE};">
      <th style="padding:6px 10px;color:#fff;border:1px solid #ddd;text-align:left;">Yard</th>
      <th style="padding:6px 10px;color:#fff;border:1px solid #ddd;text-align:center;">Issues</th>
      <th style="padding:6px 10px;color:#fff;border:1px solid #ddd;text-align:center;">NEW</th>
      <th style="padding:6px 10px;color:#fff;border:1px solid #ddd;text-align:center;">7+ Days</th>
      <th style="padding:6px 10px;color:#fff;border:1px solid #ddd;text-align:center;">In Svc</th>
      <th style="padding:6px 10px;color:#fff;border:1px solid #ddd;text-align:center;">OOS</th>
    </tr>""")

    for yard in YARD_ORDER:
        yi = grouped.get(yard, [])
        if not yi:
            continue
        y_new = len([i for i in yi if i["is_new"]])
        y_esc = len([i for i in yi if i["is_escalation"]])
        y_in = len([i for i in yi if i["availability"] == "In Service"])
        y_oos = len([i for i in yi if i["availability"] == "Out Of Service"])
        bg = "#ffffff"
        parts.append(f"""    <tr style="background:{bg};">
      <td style="padding:5px 10px;border:1px solid #ddd;font-weight:bold;">{yard}</td>
      <td style="padding:5px 10px;border:1px solid #ddd;text-align:center;">{len(yi)}</td>
      <td style="padding:5px 10px;border:1px solid #ddd;text-align:center;{f'color:#00B050;font-weight:bold;' if y_new else ''}">{y_new if y_new else '-'}</td>
      <td style="padding:5px 10px;border:1px solid #ddd;text-align:center;{f'color:#{RED};font-weight:bold;' if y_esc else ''}">{y_esc if y_esc else '-'}</td>
      <td style="padding:5px 10px;border:1px solid #ddd;text-align:center;">{y_in}</td>
      <td style="padding:5px 10px;border:1px solid #ddd;text-align:center;">{y_oos}</td>
    </tr>""")

    parts.append("  </table>\n</td></tr>")

    # ---- SECTION 2: WHAT'S NEW TODAY ----
    new_in_svc = [i for i in new_issues if i["availability"] == "In Service"]
    new_oos = [i for i in new_issues if i["availability"] != "In Service"]

    new_table_header = f"""  <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;font-size:12px;">
    <tr style="background:#{MED_BLUE};">
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Vehicle ID</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Yard</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:center;">Issue Type</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Device(s)</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Last Location</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Action</th>
    </tr>"""

    def _new_table_rows(issue_list):
        rows = ""
        for idx, issue in enumerate(issue_list):
            bg = f"#{ALT_ROW}" if idx % 2 == 1 else "#ffffff"
            issue_bg, issue_fg = _issue_html_color(issue["issue_type"])
            rows += f"""    <tr style="background:{bg};">
      <td style="padding:5px 8px;border:1px solid #ddd;font-weight:bold;">{_h(issue["vehicle_number"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(issue["yard"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;background:{issue_bg};color:{issue_fg};font-weight:bold;text-align:center;">{_h(issue["issue_type"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(issue["devices_affected"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(issue["location"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(issue["recommended_action"])}</td>
    </tr>"""
        return rows

    if new_issues:
        parts.append(f"""
<tr><td style="padding:0 40px;"><hr style="border:none;border-top:3px solid #00B050;margin:15px 0 0 0;"></td></tr>
<tr><td style="padding:15px 40px;">
  <h2 style="color:#00B050;margin:0 0 10px 0;font-size:16px;">WHAT'S NEW TODAY -- {len(new_in_svc)} In-Service{f", {len(new_oos)} OOS" if new_oos else ""}</h2>
  <div style="font-size:12px;color:#666;margin-bottom:8px;">Disconnected in the last 2 days. Just appeared on the report.</div>""")

        if new_in_svc:
            parts.append(f"""{new_table_header}
{_new_table_rows(new_in_svc)}
  </table>""")
        else:
            parts.append('  <div style="padding:10px 0;color:#666;font-style:italic;">No new in-service disconnections.</div>')

        if new_oos:
            new_oos_header = new_table_header.replace(f"#{MED_BLUE}", "#999999")
            parts.append(f"""
  <div style="background:#999999;color:#fff;padding:6px 15px;margin:15px 0 0 0;font-size:11px;font-weight:bold;">
    OUT OF SERVICE -- FOR AWARENESS ({len(new_oos)} new OOS vehicle{"s" if len(new_oos) != 1 else ""})
  </div>
{new_oos_header}
{_new_table_rows(new_oos)}
  </table>""")

        parts.append("</td></tr>")
    else:
        parts.append(f"""
<tr><td style="padding:0 40px;"><hr style="border:none;border-top:3px solid #00B050;margin:15px 0 0 0;"></td></tr>
<tr><td style="padding:15px 40px;">
  <h2 style="color:#00B050;margin:0 0 10px 0;font-size:16px;">WHAT'S NEW TODAY</h2>
  <div style="font-size:14px;color:#666;padding:10px 0;">No new disconnections in the last 2 days.</div>
</td></tr>""")

    # ---- SECTION 3: ESCALATION (7+ DAYS) ----
    esc_in_svc = [i for i in esc_issues if i["availability"] == "In Service"]
    esc_oos = [i for i in esc_issues if i["availability"] != "In Service"]

    def _esc_table_rows(esc_list):
        rows = ""
        for idx, issue in enumerate(esc_list):
            bg = f"#{ALT_ROW}" if idx % 2 == 1 else "#ffffff"
            issue_bg, issue_fg = _issue_html_color(issue["issue_type"])
            avail_style = f"background:#{OOS_BG};color:#{RED};font-weight:bold;" if issue["availability"] == "Out Of Service" else ""
            rows += f"""    <tr style="background:{bg};">
      <td style="padding:5px 8px;border:1px solid #ddd;font-weight:bold;">{_h(issue["vehicle_number"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(issue["yard"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;text-align:center;color:#{RED};font-weight:bold;">{issue["days_disconnected"]}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;{avail_style}">{_h(issue["availability"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;background:{issue_bg};color:{issue_fg};font-weight:bold;text-align:center;">{_h(issue["issue_type"])}</td>
      <td style="padding:5px 8px;border:1px solid #ddd;">{_h(issue["recommended_action"])}</td>
    </tr>"""
        return rows

    esc_table_header = f"""  <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;font-size:12px;">
    <tr style="background:#{MED_BLUE};">
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Vehicle ID</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Yard</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:center;">Days</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Availability</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:center;">Issue Type</th>
      <th style="padding:6px 8px;color:#fff;border:1px solid #ddd;text-align:left;">Action</th>
    </tr>"""

    if esc_in_svc:
        parts.append(f"""
<tr><td style="padding:0 40px;"><hr style="border:none;border-top:3px solid #{RED};margin:15px 0 0 0;"></td></tr>
<tr><td style="padding:15px 40px;">
  <h2 style="color:#{RED};margin:0 0 10px 0;font-size:16px;">ESCALATION -- {len(esc_in_svc)} In-Service Vehicle{"s" if len(esc_in_svc) != 1 else ""} Disconnected 7+ Days</h2>
  <div style="font-size:12px;color:#666;margin-bottom:8px;">These have been sitting unresolved. Safety reps and dispatch have been notified daily.</div>
{esc_table_header}
{_esc_table_rows(esc_in_svc)}
  </table>""")

        if esc_oos:
            esc_oos_header = esc_table_header.replace(f"#{MED_BLUE}", "#999999")
            parts.append(f"""
  <div style="background:#999999;color:#fff;padding:6px 15px;margin:15px 0 0 0;font-size:11px;font-weight:bold;">
    OUT OF SERVICE -- FOR AWARENESS ({len(esc_oos)} vehicle{"s" if len(esc_oos) != 1 else ""} also 7+ days)
  </div>
{esc_oos_header}
{_esc_table_rows(esc_oos)}
  </table>""")

        parts.append("</td></tr>")
    elif esc_oos:
        # No in-service escalations but some OOS
        parts.append(f"""
<tr><td style="padding:0 40px;"><hr style="border:none;border-top:3px solid #{RED};margin:15px 0 0 0;"></td></tr>
<tr><td style="padding:15px 40px;">
  <h2 style="color:#{RED};margin:0 0 10px 0;font-size:16px;">ESCALATION (7+ DAYS)</h2>
  <div style="font-size:14px;color:#666;padding:10px 0;">No in-service vehicles disconnected 7+ days.</div>
  <div style="background:#999999;color:#fff;padding:6px 15px;margin:10px 0 0 0;font-size:11px;font-weight:bold;">
    OUT OF SERVICE -- FOR AWARENESS ({len(esc_oos)} vehicle{"s" if len(esc_oos) != 1 else ""} disconnected 7+ days)
  </div>
{esc_table_header.replace(f"#{MED_BLUE}", "#999999")}
{_esc_table_rows(esc_oos)}
  </table>
</td></tr>""")
    else:
        parts.append(f"""
<tr><td style="padding:0 40px;"><hr style="border:none;border-top:3px solid #{RED};margin:15px 0 0 0;"></td></tr>
<tr><td style="padding:15px 40px;">
  <h2 style="color:#{RED};margin:0 0 10px 0;font-size:16px;">ESCALATION (7+ DAYS)</h2>
  <div style="font-size:14px;color:#666;padding:10px 0;">No vehicles have been disconnected 7+ days. All clear.</div>
</td></tr>""")

    # ---- SECTION 4: OOS ANOMALIES ----
    if oos_active_list:
        parts.append(_build_oos_anomaly_html(oos_active_list))

    # Attachment note
    parts.append(f"""
<tr><td style="padding:15px 40px;">
  <div style="background:#f8f8f8;border:1px solid #ddd;padding:12px 15px;font-size:12px;color:#666;text-align:center;">
    Full detailed report attached as Excel. Open for complete vehicle-by-vehicle breakdown by yard.
  </div>
</td></tr>""")

    parts.append(_build_html_footer("director", gen_str))
    return "\n".join(parts)


# ==============================================================================
# SEND EMAIL
# ==============================================================================

def send_email(html_body, subject, recipients_str, xlsx_path=None, cc_str=""):
    """Send report via Gmail SMTP with optional Excel attachment and CC.

    Args:
        html_body: HTML email content
        subject: Email subject line
        recipients_str: Comma-separated TO email addresses
        xlsx_path: Optional path to Excel attachment
        cc_str: Comma-separated CC email addresses
    """
    gmail_address = os.environ.get("GMAIL_ADDRESS", "")
    gmail_app_password = os.environ.get("GMAIL_APP_PASSWORD", "")

    if not gmail_address or not gmail_app_password:
        print("    Email skipped -- GMAIL_ADDRESS or GMAIL_APP_PASSWORD not set.")
        return False

    recipients = [r.strip() for r in recipients_str.split(",") if r.strip()]
    cc_list = [r.strip() for r in cc_str.split(",") if r.strip()] if cc_str else []
    if not recipients:
        return False

    try:
        msg = MIMEMultipart("mixed")
        msg["From"] = gmail_address
        msg["To"] = ", ".join(recipients)
        if cc_list:
            msg["Cc"] = ", ".join(cc_list)
        msg["Subject"] = subject

        msg.attach(MIMEText(html_body, "html"))

        if xlsx_path and os.path.exists(xlsx_path):
            with open(xlsx_path, "rb") as f:
                part = MIMEBase(
                    "application",
                    "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f'attachment; filename="{os.path.basename(xlsx_path)}"',
            )
            msg.attach(part)

        # SMTP sendmail needs all recipients (TO + CC)
        all_recipients = recipients + cc_list

        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(gmail_address, gmail_app_password)
            server.sendmail(gmail_address, all_recipients, msg.as_string())

        print(f"    Sent to: {', '.join(recipients)}")
        if cc_list:
            print(f"    CC: {', '.join(cc_list)}")
        return True
    except Exception as e:
        print(f"    Email failed: {e}")
        return False


def send_tiered_emails(issues, grouped, report_date, xlsx_path, csv_available):
    """Send emails to all configured tiers.

    Tier 1: Camera Team -- daily full report (all yards) + Excel attachment
    Tier 2: Safety Director -- daily full report + Excel attachment
    Tier 3: Safety Reps -- daily per-yard report (their yards only)
    Tier 4: Dispatch -- daily per-yard report (their yards only)
    Tier 5: Managers -- 7+ day escalation only (per-yard), only if there are escalations
    """
    date_str = report_date.strftime("%B %d, %Y")
    emails_sent = 0

    # ---- TIER 1: Camera Team (full report) ----
    cam_recipients = CAMERA_TEAM_RECIPIENTS
    if cam_recipients.strip():
        print("\n  [Camera Team] Full daily report...")
        html = build_html_email(issues, grouped, report_date, csv_available, mode="full")
        subject = f"Casing Device Status Report - {date_str}"
        if send_email(html, subject, cam_recipients, xlsx_path):
            emails_sent += 1

    # ---- TIER 2: Safety Director (30-second summary + Excel) ----
    dir_recipients = SAFETY_DIRECTOR_RECIPIENTS
    if dir_recipients.strip():
        print("\n  [Safety Director] Director summary...")
        html = build_html_email(issues, grouped, report_date, csv_available, mode="director")
        new_count = len([i for i in issues if i["is_new"]])
        esc_count = len([i for i in issues if i["is_escalation"]])
        subject = f"Casing Device Summary - {date_str}"
        tags = []
        if new_count > 0:
            tags.append(f"{new_count} NEW")
        if esc_count > 0:
            tags.append(f"{esc_count} ESCALATION")
        if tags:
            subject = f"[{' | '.join(tags)}] {subject}"
        if send_email(html, subject, dir_recipients, xlsx_path):
            emails_sent += 1

    # ---- TIER 3: Safety Reps (per-yard, their yards only) ----
    for yard in YARD_ORDER:
        rep_emails = SAFETY_REP_RECIPIENTS.get(yard, "")
        if not rep_emails.strip():
            continue
        yard_issues = grouped.get(yard, [])
        if not yard_issues:
            continue
        print(f"\n  [Safety Rep - {yard}] Yard report...")
        html = build_html_email(issues, grouped, report_date, csv_available,
                                mode="yard", yard_filter=yard)
        subject = f"{yard} Casing Device Status - {date_str}"
        if send_email(html, subject, rep_emails):
            emails_sent += 1

    # ---- TIER 4: Dispatch (per-yard, their yards only) ----
    for yard in YARD_ORDER:
        disp_emails = DISPATCH_RECIPIENTS.get(yard, "")
        if not disp_emails.strip():
            continue
        yard_issues = grouped.get(yard, [])
        if not yard_issues:
            continue
        print(f"\n  [Dispatch - {yard}] Yard report...")
        html = build_html_email(issues, grouped, report_date, csv_available,
                                mode="yard", yard_filter=yard)
        subject = f"ACTION: {yard} Device Issues - {date_str}"
        if send_email(html, subject, disp_emails):
            emails_sent += 1

    # ---- TIER 5: Managers (escalation only, 7+ days) -- CC: Buffington ----
    for yard in YARD_ORDER:
        mgr_emails = MANAGER_RECIPIENTS.get(yard, "")
        if not mgr_emails.strip():
            continue
        yard_escalations = [i for i in grouped.get(yard, []) if i["is_escalation"] and i["availability"] == "In Service"]
        if not yard_escalations:
            continue
        print(f"\n  [Manager - {yard}] Escalation alert ({len(yard_escalations)} issues)...")
        html = build_html_email(issues, grouped, report_date, csv_available,
                                mode="escalation", yard_filter=yard)
        subject = f"ESCALATION: {yard} - {len(yard_escalations)} Device(s) Disconnected 7+ Days"
        if send_email(html, subject, mgr_emails, cc_str=MANAGER_ESCALATION_CC):
            emails_sent += 1

    return emails_sent


# ==============================================================================
# MAIN
# ==============================================================================

def main():
    today = datetime.now(timezone.utc).astimezone(CENTRAL_TZ)
    report_date = today

    print("\n" + "=" * 80)
    print("DAILY CASING DEVICE & CAMERA STATUS REPORT")
    print(f"Report date: {report_date.strftime('%A, %B %d, %Y')}")
    print("=" * 80)

    # Step 1: Fetch casing vehicles
    print("\n[1] Fetching casing vehicles from Motive API...")
    vehicles = fetch_casing_vehicles()
    print(f"    {len(vehicles)} casing vehicles found")

    yard_counts = {}
    for v in vehicles.values():
        yard_counts[v["yard"]] = yard_counts.get(v["yard"], 0) + 1
    for yard in YARD_ORDER:
        if yard in yard_counts:
            print(f"      {yard}: {yard_counts[yard]}")

    # Step 2: Fetch vehicle locations
    print("\n[2] Fetching vehicle locations...")
    locations = fetch_vehicle_locations()
    casing_locs = {vid: loc for vid, loc in locations.items() if vid in vehicles}
    print(f"    {len(casing_locs)} casing vehicles with location data")

    # Step 3: Fetch Device Status CSV from Gmail
    print("\n[3] Fetching Motive reports from Gmail (IMAP)...")
    device_csv, camera_csv = fetch_csv_from_gmail()
    csv_devices = {}
    csv_available = False

    if device_csv:
        csv_devices = parse_device_csv(device_csv)
        csv_available = bool(csv_devices)
        print(f"    Device Status CSV: {len(csv_devices)} vehicles parsed")
    else:
        print("    Device Status CSV: not found")

    if camera_csv:
        print(f"    Cameras Status CSV: found (supplementary data)")
    else:
        print("    Cameras Status CSV: not found")

    if not csv_available:
        print("    Running in API-only mode (72h threshold; camera-only issues not detected)")

    # Step 4: Classify issues
    print("\n[4] Classifying device issues...")
    issues = classify_issues(vehicles, locations, csv_devices)
    print(f"    {len(issues)} vehicles with issues detected")

    powered_off = len([i for i in issues if i["issue_type"] == "Powered Off"])
    camera_off = len([i for i in issues if i["issue_type"] == "Camera Powered Off"])
    inactive = len([i for i in issues if i["issue_type"] == "Inactive 30+ Days"])
    in_svc = len([i for i in issues if i["availability"] == "In Service"])
    oos = len([i for i in issues if i["availability"] == "Out Of Service"])
    new_count = len([i for i in issues if i["is_new"]])
    esc_count = len([i for i in issues if i["is_escalation"]])
    oos_active = len([i for i in issues if i["oos_active"]])

    print(f"    Powered Off: {powered_off}")
    print(f"    Camera Powered Off: {camera_off}")
    print(f"    Inactive 30+ Days: {inactive}")
    print(f"    In Service: {in_svc} | Out of Service: {oos}")
    print(f"    NEW (< 2 days): {new_count}")
    print(f"    ESCALATION (7+ days): {esc_count}")
    if oos_active:
        print(f"    OOS with recent GPS activity: {oos_active} (status update needed)")

    # Step 4b: Fetch recent mileage for OOS-active vehicles
    oos_active_issues = [i for i in issues if i["oos_active"]]
    if oos_active_issues:
        print(f"\n[4b] Fetching recent mileage for {len(oos_active_issues)} OOS-active vehicles...")
        oos_vnums = {i["vehicle_number"] for i in oos_active_issues}
        oos_miles = fetch_oos_mileage(oos_vnums, days=7)
        for issue in issues:
            if issue["oos_active"]:
                issue["recent_miles"] = oos_miles.get(issue["vehicle_number"], 0)
        for vnum, miles in sorted(oos_miles.items(), key=lambda x: x[1], reverse=True):
            if miles > 0:
                print(f"      {vnum}: {miles} miles (last 7 days)")
    else:
        for issue in issues:
            issue["recent_miles"] = 0

    # Step 5: Group by yard
    print("\n[5] Grouping by yard...")
    grouped = group_issues_by_yard(issues)
    for yard, yard_issues in grouped.items():
        y_new = len([i for i in yard_issues if i["is_new"]])
        y_esc = len([i for i in yard_issues if i["is_escalation"]])
        extra = ""
        if y_new:
            extra += f" ({y_new} NEW)"
        if y_esc:
            extra += f" ({y_esc} ESCALATION)"
        print(f"    {yard}: {len(yard_issues)} issue{'s' if len(yard_issues) != 1 else ''}{extra}")

    # Step 6: Generate Excel
    print("\n[6] Generating Excel report...")
    wb = create_excel_report(issues, grouped, report_date)
    date_tag = report_date.strftime("%Y-%m-%d")
    xlsx_filename = f"Casing_Device_Status_Report_{date_tag}.xlsx"
    wb.save(xlsx_filename)
    print(f"    Saved: {xlsx_filename}")

    # Step 7: Send tiered emails
    print("\n[7] Sending tiered emails...")
    emails_sent = send_tiered_emails(issues, grouped, report_date, xlsx_filename, csv_available)
    print(f"\n    Total emails sent: {emails_sent}")

    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80 + "\n")


if __name__ == "__main__":
    main()
