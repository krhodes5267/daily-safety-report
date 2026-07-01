"""
Universal Safety Recap -- Data Fetcher.

Centralized data collection from KPA API, Motive API, and man-hours Excel.
All API patterns reused from existing scripts (casing_monthly_recap.py,
daily_speeding_report.py, casing_field_assessment_audit.py).
"""

import csv
import json
import os
import time
from datetime import datetime, timezone
from io import StringIO

import requests

from .config import (
    DIVISION_LOB_IDS,
    KPA_BASE_URL,
    KPA_CALL_DELAY,
    KPA_ENDPOINTS,
    LOB_ID_MAP,
    MAN_HOURS_CO_CODES,
    MAN_HOURS_DEPARTMENTS,
    MOTIVE_BASE_V1,
    MOTIVE_BASE_V2,
    MOTIVE_GROUP_MAP,
    SHARED_FORMS,
    normalize_location,
)

try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo

CENTRAL_TZ = ZoneInfo("America/Chicago")
KMH_TO_MPH = 0.621371


# ===========================================================================
# KPA API Client
# ===========================================================================

class KPAClient:
    """KPA EHS API client with POST-based auth, rate-limit retry, CSV pagination."""

    def __init__(self, token=None):
        self.token = token or os.environ.get("KPA_API_TOKEN", "")
        if not self.token:
            raise ValueError("KPA_API_TOKEN not set")

    def _call(self, endpoint, params):
        """Single KPA API call with rate-limit retry."""
        url = f"{KPA_BASE_URL}/{endpoint}"
        payload = {"token": self.token, "limit": 1000}
        payload.update(params)

        for attempt in range(3):
            try:
                r = requests.post(url, json=payload, timeout=60)
                text = r.text.strip()
                if "rate_limit" in text:
                    wait = 30 * (attempt + 1)
                    print(f"    Rate limited -- waiting {wait}s...")
                    time.sleep(wait)
                    continue
                return text if text else None
            except Exception as e:
                print(f"    KPA API error ({endpoint}): {e}")
                if attempt < 2:
                    time.sleep(5)
                    continue
                return None
        return None

    def _call_paginated(self, endpoint, params):
        """Paginated KPA call returning all CSV rows as list of dicts."""
        all_rows = []
        headers = None
        page = 1

        while True:
            p = dict(params)
            p["page"] = page
            text = self._call(endpoint, p)
            time.sleep(KPA_CALL_DELAY)

            if not text:
                break
            try:
                reader = csv.DictReader(StringIO(text))
                rows = list(reader)
                # Filter out header echo rows
                data = [r for r in rows if r.get("date", "") != "Date"]
                if not data:
                    break
                if headers is None:
                    headers = reader.fieldnames
                all_rows.extend(data)
                if len(rows) < 1000:
                    break
                page += 1
            except Exception:
                break

        return headers or [], all_rows

    # --- High-level data fetchers ---

    def get_form_responses(self, form_id, start_date, end_date, extra_params=None):
        """Fetch form responses for a date range. Returns (headers, rows).

        KPA responses.flat uses 'updated_after' (epoch ms) instead of start_date/end_date,
        and requires 'format': 'csv'. Rows are then filtered by the 'date' field to stay
        within the requested date range.
        """
        # Convert start_date to epoch milliseconds for updated_after
        start_ms = self._date_to_epoch_ms(start_date)

        params = {
            "form_id": form_id,
            "format": "csv",
            "updated_after": start_ms,
        }
        if extra_params:
            params.update(extra_params)

        headers, rows = self._call_paginated(KPA_ENDPOINTS["responses"], params)

        # Filter rows to the requested date range (updated_after may return older edits).
        # Prefer the "Date Conducted" form field (tm4zqob5uficucju) over the submission
        # timestamp (date), since assessors may submit days after the field visit.
        ASSESSMENT_DATE_HASH = "tm4zqob5uficucju"
        filtered = []
        for row in rows:
            # Try assessment date field first (format: "5/25/2026 6:30 AM")
            conducted = row.get(ASSESSMENT_DATE_HASH, "").strip()
            if conducted:
                try:
                    for fmt in ("%m/%d/%Y %I:%M %p", "%m/%d/%Y"):
                        try:
                            dt = datetime.strptime(conducted, fmt)
                            conducted_iso = dt.strftime("%Y-%m-%d")
                            break
                        except ValueError:
                            continue
                    else:
                        conducted_iso = None
                    if conducted_iso and start_date <= conducted_iso <= end_date:
                        filtered.append(row)
                        continue
                    elif conducted_iso:
                        continue  # Has a conducted date but outside range
                except Exception:
                    pass  # Fall through to submission date check

            row_date = row.get("date", "")
            if row_date and start_date <= row_date[:10] <= end_date:
                filtered.append(row)
            elif not row_date:
                # Include rows without a date field (some forms use different date keys)
                filtered.append(row)

        return headers, filtered

    @staticmethod
    def _date_to_epoch_ms(date_str):
        """Convert 'YYYY-MM-DD' to epoch milliseconds."""
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return int(dt.replace(tzinfo=timezone.utc).timestamp() * 1000)

    # --- JSON endpoints (followups, training, users) ---

    def _call_json_paginated(self, endpoint, params, data_key, max_pages=50):
        """Paginated KPA call for JSON endpoints. Returns (field_names, rows_as_dicts).

        JSON endpoints (followups, training, users) support max limit of 500,
        unlike responses.flat which supports 1000.
        max_pages caps pagination to avoid very long fetches (500 * 50 = 25,000 rows).
        """
        all_rows = []
        page = 1

        while True:
            p = dict(params)
            p["page"] = page
            url = f"{KPA_BASE_URL}/{endpoint}"
            payload = {"token": self.token, "limit": 500}
            payload.update(p)

            for attempt in range(3):
                try:
                    r = requests.post(url, json=payload, timeout=60)
                    text = r.text.strip()
                    if "rate_limit" in text:
                        wait = 30 * (attempt + 1)
                        print(f"    Rate limited -- waiting {wait}s...")
                        time.sleep(wait)
                        continue
                    break
                except Exception as e:
                    print(f"    KPA JSON API error ({endpoint}): {e}")
                    if attempt < 2:
                        time.sleep(5)
                    text = None
                    break

            time.sleep(KPA_CALL_DELAY)
            if not text:
                break

            try:
                data = json.loads(text)
            except Exception:
                break

            items = data.get(data_key, [])
            if not items:
                break
            all_rows.extend(items)

            paging = data.get("paging", {})
            last_page = paging.get("last_page", 1)
            if page >= last_page:
                break
            if page >= max_pages:
                print(f"    Capped at {max_pages} pages ({len(all_rows)} rows)")
                break
            page += 1

        field_names = list(all_rows[0].keys()) if all_rows else []
        return field_names, all_rows

    def get_followups(self):
        """Fetch ALL corrective actions / follow-ups (JSON endpoint).

        Returns (field_names, rows) where each row is a dict with keys:
        form_id, response_id, id, open, due (YYYYMMDD int), created_on,
        updated_on, resolved_on, m_observer_id, m_assigner_id, m_assignee_id,
        m_completer_id, messages.

        Due dates come as integers like 20260410 (April 10, 2026).
        """
        fields, rows = self._call_json_paginated(
            KPA_ENDPOINTS["followups"], {}, "followups"
        )
        # Normalize: convert due (int 20260410) to ISO string, add status field
        for row in rows:
            due_int = row.get("due")
            if due_int and isinstance(due_int, int):
                row["due_date"] = f"{str(due_int)[:4]}-{str(due_int)[4:6]}-{str(due_int)[6:8]}"
            else:
                row["due_date"] = ""
            row["status"] = "Open" if row.get("open") else "Closed"
            # Extract the first message note as the finding description
            msgs = row.get("messages", [])
            if msgs and isinstance(msgs, list):
                row["finding"] = msgs[0].get("note", "") or ""
            else:
                row["finding"] = ""
        return fields, rows

    def get_completed_trainings(self, start_date=None, end_date=None):
        """Fetch training completion records (JSON endpoint).

        NOTE: This endpoint has 157K+ records with no date filter param.
        We cap at 10 pages (5,000 records) to keep fetch time reasonable.
        For monthly reports, the training-employee-status endpoint provides
        the compliance data we actually need.
        """
        fields, rows = self._call_json_paginated(
            KPA_ENDPOINTS["completed_trainings"], {}, "completedtrainings",
            max_pages=10,
        )
        # Filter by date_number if date range specified
        if start_date and end_date:
            start_int = int(start_date.replace("-", ""))
            end_int = int(end_date.replace("-", ""))
            rows = [r for r in rows if start_int <= (r.get("date_number") or 0) <= end_int]
        return fields, rows

    def get_training_employee_status(self):
        """Fetch training compliance status per employee (JSON endpoint).

        Returns (field_names, rows). Each row has: m_user_id, status,
        incomplete_training_ids, complete_training_ids, percent_complete.
        """
        return self._call_json_paginated(
            KPA_ENDPOINTS["training_status"], {}, "employees"
        )

    def get_training_programs(self):
        """Fetch training program catalog (JSON endpoint)."""
        return self._call_json_paginated(
            KPA_ENDPOINTS["trainings"], {}, "trainings"
        )

    def get_users(self, status="active"):
        """Fetch employee roster (JSON endpoint)."""
        return self._call_json_paginated(
            KPA_ENDPOINTS["users"], {"status": status}, "users"
        )


# ===========================================================================
# Motive API Client
# ===========================================================================

class MotiveClient:
    """Motive (formerly KeepTruckin) API client for fleet data."""

    def __init__(self, api_key=None):
        self.api_key = api_key or os.environ.get("MOTIVE_API_KEY", "")
        if not self.api_key:
            raise ValueError("MOTIVE_API_KEY not set")
        self.headers = {"X-Api-Key": self.api_key}

    def get_vehicles(self):
        """Fetch all vehicles. Returns dict: {vehicle_number: vehicle_data}."""
        vehicles = {}
        page = 1

        while True:
            try:
                resp = requests.get(
                    f"{MOTIVE_BASE_V1}/vehicles",
                    headers=self.headers,
                    params={"per_page": 100, "page_no": page},
                    timeout=30,
                )
                resp.raise_for_status()
                data = resp.json()
                vlist = data.get("vehicles", [])
                if not vlist:
                    break

                for wrapper in vlist:
                    v = wrapper.get("vehicle", wrapper)
                    num = v.get("number", "")
                    if not num:
                        continue

                    # Driver name
                    driver_name = None
                    for field in ("current_driver", "permanent_driver"):
                        d = v.get(field)
                        if d and isinstance(d, dict):
                            name = f"{d.get('first_name', '')} {d.get('last_name', '')}".strip()
                            if name:
                                driver_name = name
                                break

                    # Division/yard from group IDs
                    group_ids = v.get("group_ids", [])
                    division = None
                    yard = None
                    for gid in group_ids:
                        if gid in MOTIVE_GROUP_MAP:
                            division, yard = MOTIVE_GROUP_MAP[gid]
                            break

                    vehicles[num] = {
                        "number": num,
                        "driver": driver_name or "Unknown",
                        "division": division,
                        "yard": yard,
                        "group_ids": group_ids,
                        "raw": v,
                    }

                pag = data.get("pagination", {})
                if page * 100 >= pag.get("total", 0):
                    break
                page += 1

            except Exception as e:
                print(f"    Warning: vehicle page {page} failed: {e}")
                break

        return vehicles

    def get_ifta_trips(self, start_date, end_date, vehicle_ids=None):
        """Fetch IFTA trip data for mileage. Returns list of trip dicts.

        Endpoint: GET /v1/ifta/trips (not /trip_reports).
        Response: {"ifta_trips": [{"ifta_trip": {...}}, ...]}
        Each trip has: vehicle.number, distance (miles), jurisdiction, etc.
        """
        trips = []
        page = 1

        while True:
            params = {
                "per_page": 100,
                "page_no": page,
                "start_date": start_date,
                "end_date": end_date,
            }
            if vehicle_ids:
                params["vehicle_ids"] = ",".join(str(v) for v in vehicle_ids)

            try:
                resp = requests.get(
                    f"{MOTIVE_BASE_V1}/ifta/trips",
                    headers=self.headers,
                    params=params,
                    timeout=60,
                )
                resp.raise_for_status()
                data = resp.json()
                trip_list = data.get("ifta_trips", [])
                if not trip_list:
                    break

                # Unwrap {"ifta_trip": {...}} wrappers
                for wrapper in trip_list:
                    trip = wrapper.get("ifta_trip", wrapper)
                    # Normalize: ensure total_miles field exists for downstream code
                    if "total_miles" not in trip and "distance" in trip:
                        trip["total_miles"] = trip["distance"]
                    trips.append(trip)

                pag = data.get("pagination", {})
                if page * 100 >= pag.get("total", 0):
                    break
                page += 1
            except Exception as e:
                print(f"    Warning: IFTA page {page} failed: {e}")
                break

        return trips

    def get_speeding_events(self, start_date, end_date):
        """Fetch V1 speeding events. Returns list of event dicts."""
        # Convert dates to UTC boundaries
        start_central = datetime(
            start_date.year, start_date.month, start_date.day,
            0, 0, 0, tzinfo=CENTRAL_TZ,
        )
        end_central = datetime(
            end_date.year, end_date.month, end_date.day,
            23, 59, 59, tzinfo=CENTRAL_TZ,
        )
        start_utc = start_central.astimezone(timezone.utc)
        end_utc = end_central.astimezone(timezone.utc)

        api_start = start_utc.strftime("%Y-%m-%d")
        api_end = end_utc.strftime("%Y-%m-%d")

        events = []
        page = 1

        while True:
            params = {
                "per_page": 100,
                "page_no": page,
                "start_date": api_start,
                "end_date": api_end,
            }
            try:
                resp = requests.get(
                    f"{MOTIVE_BASE_V1}/speeding_events",
                    headers=self.headers,
                    params=params,
                    timeout=30,
                )
                resp.raise_for_status()
                data = resp.json()
                evts = data.get("speeding_events", [])
                if not evts:
                    break
                events.extend(evts)

                total = data.get("total", 0)
                if page * 100 >= total:
                    break
                page += 1
            except Exception as e:
                print(f"    Warning: speeding page {page} failed: {e}")
                break

        # Filter to Central Time window
        filtered = []
        for wrapper in events:
            evt = wrapper.get("speeding_event", wrapper)
            evt_time_str = evt.get("start_time", "")
            try:
                evt_utc = datetime.fromisoformat(evt_time_str.replace("Z", "+00:00"))
                evt_central = evt_utc.astimezone(CENTRAL_TZ)
                if start_central <= evt_central <= end_central:
                    filtered.append(evt)
            except Exception:
                filtered.append(evt)

        return filtered

    def get_camera_events(self, start_date, end_date):
        """Fetch V2 driver performance (camera) events. Page-number paginated."""
        start_central = datetime(
            start_date.year, start_date.month, start_date.day,
            0, 0, 0, tzinfo=CENTRAL_TZ,
        )
        end_central = datetime(
            end_date.year, end_date.month, end_date.day,
            23, 59, 59, tzinfo=CENTRAL_TZ,
        )
        start_utc = start_central.astimezone(timezone.utc)
        end_utc = end_central.astimezone(timezone.utc)

        events = []
        page = 1

        while True:
            params = {
                "per_page": 100,
                "page_no": page,
                "start_date": start_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
                "end_date": end_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
            }

            try:
                resp = requests.get(
                    f"{MOTIVE_BASE_V2}/driver_performance_events",
                    headers=self.headers,
                    params=params,
                    timeout=30,
                )
                resp.raise_for_status()
                data = resp.json()
                evts = data.get("driver_performance_events", [])
                if not evts:
                    break
                # Unwrap {"driver_performance_event": {...}} wrappers
                for wrapper in evts:
                    evt = wrapper.get("driver_performance_event", wrapper)
                    events.append(evt)

                if len(evts) < 100:
                    break
                page += 1
            except Exception as e:
                print(f"    Warning: camera events page {page} failed: {e}")
                break

        return events


# ===========================================================================
# Man-Hours Excel Parser
# ===========================================================================

class ManHoursParser:
    """Parse man-hours Excel file for TRIR/DART calculations."""

    def __init__(self, file_path=None):
        self.file_path = file_path or os.path.join(
            os.path.expanduser("~"), "Downloads",
            "2025 Q1 Man Hours Division Breakdown.xlsx",
        )

    def parse(self, sheet_name, division_filter=None):
        """Parse a sheet from the man-hours workbook.

        Args:
            sheet_name: Sheet name in Excel (BRHAS, BTI, Transcend, etc.)
            division_filter: If set, only include rows where Division Name matches.

        Returns:
            dict with keys:
                total_regular: float
                total_overtime: float
                total_hours: float
                by_location: {location: {regular, overtime, total}}
                by_employee: [{name, division, department, location, regular, overtime, total, status}]
        """
        try:
            import openpyxl
        except ImportError:
            print("    Warning: openpyxl not installed -- man-hours data unavailable")
            return self._empty_result()

        if not os.path.exists(self.file_path):
            print(f"    Warning: Man-hours file not found: {self.file_path}")
            return self._empty_result()

        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            print(f"    Warning: Could not open man-hours file: {e}")
            return self._empty_result()

        if sheet_name not in wb.sheetnames:
            print(f"    Warning: Sheet '{sheet_name}' not found in man-hours file")
            return self._empty_result()

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return self._empty_result()

        # Find header row
        header_idx = None
        for i, row in enumerate(rows):
            cells = [str(c).strip().lower() if c else "" for c in row]
            if "first name" in cells or "last name" in cells:
                header_idx = i
                break

        if header_idx is None:
            print(f"    Warning: Could not find header row in sheet '{sheet_name}'")
            return self._empty_result()

        headers = [str(c).strip() if c else "" for c in rows[header_idx]]
        col = {h.lower().rstrip("\t"): i for i, h in enumerate(headers)}

        result = {
            "total_regular": 0.0,
            "total_overtime": 0.0,
            "total_hours": 0.0,
            "by_location": {},
            "by_employee": [],
        }

        name_col = col.get("name", col.get("id"))
        div_col = col.get("division name")
        dept_col = col.get("department name")
        loc_col = col.get("location name")
        reg_col = col.get("regular hours", col.get("regular"))
        ot_col = col.get("overtime hours", col.get("overtime"))
        status_col = col.get("emp status descr", col.get("emp status"))
        first_col = col.get("first name")
        last_col = col.get("last name")

        for row in rows[header_idx + 1:]:
            # Skip subtotal rows (Name is None/empty)
            if name_col is not None and (row[name_col] is None or str(row[name_col]).strip() == ""):
                continue

            # Get division name, strip trailing tabs
            div_name = ""
            if div_col is not None and row[div_col]:
                div_name = str(row[div_col]).strip().rstrip("\t")

            # Apply division filter
            if division_filter and div_name.lower() != division_filter.lower():
                continue

            # Parse hours
            regular = self._parse_hours(row[reg_col] if reg_col is not None else None)
            overtime = self._parse_hours(row[ot_col] if ot_col is not None else None)
            total = regular + overtime

            # Location (normalized to match yard names)
            location = ""
            if loc_col is not None and row[loc_col]:
                location = normalize_location(str(row[loc_col]).strip())

            # Employee name
            first = str(row[first_col]).strip() if first_col is not None and row[first_col] else ""
            last = str(row[last_col]).strip() if last_col is not None and row[last_col] else ""
            emp_name = f"{first} {last}".strip()

            # Status
            status = ""
            if status_col is not None and row[status_col]:
                status = str(row[status_col]).strip()

            result["total_regular"] += regular
            result["total_overtime"] += overtime
            result["total_hours"] += total

            if location:
                if location not in result["by_location"]:
                    result["by_location"][location] = {"regular": 0, "overtime": 0, "total": 0}
                result["by_location"][location]["regular"] += regular
                result["by_location"][location]["overtime"] += overtime
                result["by_location"][location]["total"] += total

            result["by_employee"].append({
                "name": emp_name,
                "division": div_name,
                "department": str(row[dept_col]).strip().rstrip("\t") if dept_col is not None and row[dept_col] else "",
                "location": location,
                "regular": regular,
                "overtime": overtime,
                "total": total,
                "status": status,
            })

        wb.close()
        return result

    def parse_2026(self, month_str, division_key=None):
        """Parse 2026-format man-hours file (single Detail sheet, all companies).

        Args:
            month_str: Month in 'YYYY-MM' format (e.g., '2026-03')
            division_key: Division key for department/co-code filtering

        Returns:
            Same dict structure as parse(): {total_regular, total_overtime,
            total_hours, by_location, by_employee}
        """
        try:
            import openpyxl
        except ImportError:
            print("    Warning: openpyxl not installed -- man-hours data unavailable")
            return self._empty_result()

        if not os.path.exists(self.file_path):
            print(f"    Warning: Man-hours file not found: {self.file_path}")
            return self._empty_result()

        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            print(f"    Warning: Could not open man-hours file: {e}")
            return self._empty_result()

        if "Detail" not in wb.sheetnames:
            print(f"    Warning: 'Detail' sheet not found in man-hours file")
            wb.close()
            return self._empty_result()

        ws = wb["Detail"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return self._empty_result()

        # Parse header row
        headers = [str(c).strip() if c else "" for c in rows[0]]
        col = {h: i for i, h in enumerate(headers)}

        # Required columns
        month_col = col.get("Month/Year")
        name_col = col.get("Name")
        dept_col = col.get("Home Department")
        loc_col = col.get("Home Location")
        hours_col = col.get("Total Hours")
        co_col = col.get("Co Code")
        status_col = col.get("Employee Status")

        if month_col is None or hours_col is None:
            print("    Warning: Required columns not found in Detail sheet")
            wb.close()
            return self._empty_result()

        # Build month filter string (e.g., "03/2026")
        parts = month_str.split("-")
        month_filter = f"{parts[1]}/{parts[0]}"

        # Build department filter
        dept_filter = set()
        co_filter = set()
        if division_key:
            dept_filter = set(MAN_HOURS_DEPARTMENTS.get(division_key, []))
            co_filter = set(MAN_HOURS_CO_CODES.get(division_key, []))

        result = {
            "total_regular": 0.0,
            "total_overtime": 0.0,
            "total_hours": 0.0,
            "by_location": {},
            "by_employee": [],
            "headcount": 0,
        }

        for row in rows[1:]:
            # Filter by month
            row_month = str(row[month_col]).strip() if row[month_col] else ""
            if row_month != month_filter:
                continue

            # Filter by department or co code
            row_dept = str(row[dept_col]).strip() if dept_col is not None and row[dept_col] else ""
            row_co = str(row[co_col]).strip() if co_col is not None and row[co_col] else ""

            if dept_filter or co_filter:
                dept_match = row_dept in dept_filter if dept_filter else False
                co_match = row_co in co_filter if co_filter else False
                if not dept_match and not co_match:
                    continue

            # Skip inactive employees
            row_status = str(row[status_col]).strip() if status_col is not None and row[status_col] else ""
            if row_status and row_status.upper() not in ("A", "ACTIVE"):
                continue

            # Parse hours
            total = self._parse_hours(row[hours_col] if hours_col is not None else None)

            # Location
            location = ""
            if loc_col is not None and row[loc_col]:
                location = normalize_location(str(row[loc_col]).strip())

            # Employee name
            emp_name = str(row[name_col]).strip() if name_col is not None and row[name_col] else ""
            # Strip co code/emp ID from name (e.g., "Acevedo, Abi (55BRH05-BR7439)")
            if "(" in emp_name:
                emp_name = emp_name[:emp_name.index("(")].strip()

            result["total_hours"] += total
            result["headcount"] += 1

            if location:
                if location not in result["by_location"]:
                    result["by_location"][location] = {"regular": 0, "overtime": 0, "total": 0}
                result["by_location"][location]["total"] += total

            result["by_employee"].append({
                "name": emp_name,
                "division": row_dept,
                "department": row_dept,
                "location": location,
                "regular": 0,
                "overtime": 0,
                "total": total,
                "status": row_status,
            })

        wb.close()
        print(f"    Man-hours: {result['headcount']} employees, {result['total_hours']:,.1f} hours")
        return result

    def has_detail_sheet(self):
        """Check if the man-hours file has a 2026-style 'Detail' sheet."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            has_it = "Detail" in wb.sheetnames
            wb.close()
            return has_it
        except Exception:
            return False

    @staticmethod
    def _parse_hours(val):
        """Parse hours value, handling None/empty/string."""
        if val is None:
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _empty_result():
        return {
            "total_regular": 0.0,
            "total_overtime": 0.0,
            "total_hours": 0.0,
            "by_location": {},
            "by_employee": [],
        }


# ===========================================================================
# Unified Data Collector
# ===========================================================================

class DataCollector:
    """Orchestrates data collection from all sources for a division + period."""

    def __init__(self, division_config, kpa_token=None, motive_key=None, man_hours_path=None):
        self.config = division_config
        self.kpa = KPAClient(token=kpa_token)
        self.motive = MotiveClient(api_key=motive_key) if division_config.get("motive_group_ids") else None
        self.man_hours = ManHoursParser(file_path=man_hours_path)

        # Cached data
        self._vehicles = None
        self._division_user_ids = None
        self._user_lookup = None       # {user_id: "First Last"}
        self._training_lookup = None   # {training_id: "Training Name"}

    def _get_division_user_ids(self):
        """Fetch all users and return set of user IDs belonging to this division.

        Uses lineOfBusiness_id on each KPA user to match against the division's
        LOB IDs from DIVISION_LOB_IDS config.
        """
        if self._division_user_ids is not None:
            return self._division_user_ids

        # Determine which LOB IDs belong to this division
        # First check config for explicit division key
        div_key = None
        from .config import DIVISIONS
        for k, v in DIVISIONS.items():
            if v is self.config:
                div_key = k
                break

        if not div_key:
            self._division_user_ids = set()
            return self._division_user_ids

        target_lob_ids = DIVISION_LOB_IDS.get(div_key, set())
        target_fo_ids = set(self.config.get("kpa_field_office_ids", []))

        if not target_lob_ids and not target_fo_ids:
            self._division_user_ids = set()
            return self._division_user_ids

        # Fetch all users (no pagination needed -- single response)
        import requests as _req
        url = f"{KPA_BASE_URL}/users.list"
        payload = {"token": self.kpa.token}
        try:
            r = _req.post(url, json=payload, timeout=120)
            data = json.loads(r.text)
            users = data.get("users", [])
        except Exception as e:
            print(f"    Warning: Could not fetch users for division filtering: {e}")
            self._division_user_ids = set()
            return self._division_user_ids

        matched = set()
        for u in users:
            # Skip terminated employees
            if u.get("terminationDate"):
                continue
            uid = u.get("id", "")
            if not uid:
                continue

            # Match by field office if configured (e.g., Valor)
            if target_fo_ids:
                fo = u.get("fieldOffice_id", "")
                if isinstance(fo, list):
                    fo_match = any(f in target_fo_ids for f in fo)
                else:
                    fo_match = fo in target_fo_ids
                if fo_match:
                    matched.add(uid)
                    continue

            # Match by LOB
            if target_lob_ids:
                user_lobs = u.get("lineOfBusiness_id", [])
                if not isinstance(user_lobs, list):
                    user_lobs = [user_lobs]
                if target_lob_ids.intersection(user_lobs):
                    matched.add(uid)

        self._division_user_ids = matched
        print(f"    Division users: {len(matched)} (from {len(users)} total)")
        return self._division_user_ids

    def _get_user_lookup(self):
        """Build {user_id: 'First Last'} lookup from KPA users.list.

        Reuses the same API call as _get_division_user_ids when possible.
        """
        if self._user_lookup is not None:
            return self._user_lookup

        import requests as _req
        url = f"{KPA_BASE_URL}/users.list"
        payload = {"token": self.kpa.token}
        try:
            r = _req.post(url, json=payload, timeout=120)
            data = json.loads(r.text)
            users = data.get("users", [])
        except Exception as e:
            print(f"    Warning: Could not fetch users for name lookup: {e}")
            self._user_lookup = {}
            return self._user_lookup

        lookup = {}
        for u in users:
            uid = u.get("id", "")
            first = u.get("firstname", "")
            last = u.get("lastname", "")
            if uid and (first or last):
                lookup[uid] = f"{first} {last}".strip()
        self._user_lookup = lookup
        return self._user_lookup

    def _get_training_lookup(self):
        """Build {training_id: 'Training Name'} from trainings.v2.list.

        Training IDs are integers (e.g. 36472). The endpoint does not support
        limit/page params -- it returns all trainings in one call.
        Also stores creation timestamps for date filtering.
        """
        if self._training_lookup is not None:
            return self._training_lookup

        print("    Fetching training programs...")
        import requests as _req
        url = f"{KPA_BASE_URL}/{KPA_ENDPOINTS['trainings']}"
        payload = {"token": self.kpa.token}
        try:
            r = _req.post(url, json=payload, timeout=120)
            data = json.loads(r.text)
            programs = data.get("trainings", [])
        except Exception as e:
            print(f"    Warning: Could not fetch training programs: {e}")
            self._training_lookup = {}
            self._training_created = {}
            return self._training_lookup

        lookup = {}
        created = {}
        for p in programs:
            tid = p.get("id")
            name = p.get("title", p.get("name", ""))
            if tid is not None and name:
                lookup[tid] = name
            # Store creation timestamp (ms epoch)
            ts = p.get("created", 0)
            if tid is not None and ts:
                created[tid] = ts
        self._training_lookup = lookup
        self._training_created = created
        print(f"    Training programs: {len(lookup)}")
        return self._training_lookup

    def get_vehicles(self):
        """Fetch and cache Motive vehicles, filtered to this division."""
        if self._vehicles is None and self.motive:
            all_vehicles = self.motive.get_vehicles()
            group_ids = set(self.config.get("motive_group_ids", []))
            self._vehicles = {}
            for num, v in all_vehicles.items():
                if group_ids.intersection(v.get("group_ids", [])):
                    self._vehicles[num] = v
        return self._vehicles or {}

    def collect_monthly(self, year, month):
        """Collect all data for a monthly report.

        Returns dict with keys matching section names, each containing
        the raw data needed by that section's renderer.
        """
        from calendar import monthrange
        last_day = monthrange(year, month)[1]
        start = f"{year}-{month:02d}-01"
        end = f"{year}-{month:02d}-{last_day:02d}"

        # Previous month for MoM comparison
        if month == 1:
            prev_year, prev_month = year - 1, 12
        else:
            prev_year, prev_month = year, month - 1
        prev_last = monthrange(prev_year, prev_month)[1]
        prev_start = f"{prev_year}-{prev_month:02d}-01"
        prev_end = f"{prev_year}-{prev_month:02d}-{prev_last:02d}"

        data = {
            "division": self.config,
            "year": year,
            "month": month,
            "period": "monthly",
            "start_date": start,
            "end_date": end,
            "prev_start": prev_start,
            "prev_end": prev_end,
            # Ensure all expected keys exist with empty defaults
            "incidents_current": ([], []),
            "incidents_prev": ([], []),
            "observations_current": ([], []),
            "observations_prev": ([], []),
            "hse_observations_current": ([], []),
            "hse_observations_prev": ([], []),
            "training_status": ([], []),
            "completed_trainings": ([], []),
            "followups": ([], []),
            "assessments_current": ([], []),
            "assessments_prev": ([], []),
            "jsas_current": ([], []),
            "jsas_prev": ([], []),
            "jsa_reviews_current": ([], []),
            "rig_inspections_current": ([], []),
            "rig_inspections_prev": ([], []),
            "vehicle_inspections_current": ([], []),
            "vehicle_inspections_prev": ([], []),
            "man_hours": {"total_hours": 0, "by_location": {}, "by_employee": []},
            "ifta_trips": [],
            "speeding_events": [],
            "camera_events": [],
            "vehicles": {},
        }

        enabled = {s for s, on in self.config.get("sections", {}).items() if on}

        print(f"\n  Collecting data for {self.config['display_name']} ({start} to {end})...")

        # -- KPA data --
        if enabled.intersection({"incidents", "executive_summary"}):
            print("    Fetching incidents...")
            data["incidents_current"] = self._fetch_form(SHARED_FORMS["incident"], start, end)
            data["incidents_prev"] = self._fetch_form(SHARED_FORMS["incident"], prev_start, prev_end)

        if enabled.intersection({"observations", "executive_summary"}):
            print("    Fetching observations...")
            obs_form = self.config.get("form_ids", {}).get("observation", SHARED_FORMS["observation"])
            data["observations_current"] = self._fetch_form(obs_form, start, end)
            data["observations_prev"] = self._fetch_form(obs_form, prev_start, prev_end)
            # Also fetch shared HSE obs if division has its own form
            if obs_form != SHARED_FORMS["observation"]:
                data["hse_observations_current"] = self._fetch_form(SHARED_FORMS["observation"], start, end)
                data["hse_observations_prev"] = self._fetch_form(SHARED_FORMS["observation"], prev_start, prev_end)

        if "training" in enabled:
            print("    Fetching training status...")
            data["training_status"] = self._fetch_training_status(end_date=end)
            data["completed_trainings"] = self._fetch_completed_trainings(start, end)

        if "corrective_actions" in enabled:
            print("    Fetching corrective actions...")
            data["followups"] = self._fetch_followups()

        if "assessments" in enabled:
            print("    Fetching safety assessments...")
            data["assessments_current"] = self._fetch_assessments(start, end)
            data["assessments_prev"] = self._fetch_assessments(prev_start, prev_end)

        if "rig_inspections" in enabled:
            rig_insp_form = self.config.get("form_ids", {}).get("rig_inspection")
            if rig_insp_form:
                print("    Fetching rig inspections...")
                # Division-specific form -- skip company/service_line filtering
                data["rig_inspections_current"] = self.kpa.get_form_responses(rig_insp_form, start, end)
                data["rig_inspections_prev"] = self.kpa.get_form_responses(rig_insp_form, prev_start, prev_end)

        if "jsas" in enabled:
            print("    Fetching JSAs...")
            jsa_form = self.config.get("form_ids", {}).get("jsa", SHARED_FORMS["jsa_log"])
            data["jsas_current"] = self._fetch_form(jsa_form, start, end)
            data["jsas_prev"] = self._fetch_form(jsa_form, prev_start, prev_end)
            # JSA reviews
            jsa_review = self.config.get("form_ids", {}).get("jsa_review", SHARED_FORMS.get("jsa_review"))
            if jsa_review:
                data["jsa_reviews_current"] = self._fetch_form(jsa_review, start, end)

        if enabled.intersection({"vehicle_inspections", "equipment_inspections"}):
            print("    Fetching inspections...")
            data["vehicle_inspections_current"] = self._fetch_form(SHARED_FORMS["vehicle_inspection"], start, end)
            data["vehicle_inspections_prev"] = self._fetch_form(SHARED_FORMS["vehicle_inspection"], prev_start, prev_end)

        # -- Motive data --
        if self.motive and enabled.intersection({"fleet_mileage", "speeding", "camera_events"}):
            from datetime import date as date_cls
            start_dt = date_cls(year, month, 1)
            end_dt = date_cls(year, month, last_day)

            # Get division vehicles first for filtering
            data["vehicles"] = self.get_vehicles()
            div_vehicle_nums = set(data["vehicles"].keys())
            print(f"    Division vehicles: {len(div_vehicle_nums)}")

            if "fleet_mileage" in enabled:
                print("    Fetching IFTA trips...")
                all_trips = self.motive.get_ifta_trips(start, end)
                # Filter to division vehicles
                data["ifta_trips"] = self._filter_trips_to_vehicles(all_trips, div_vehicle_nums)
                print(f"    IFTA: {len(all_trips)} total -> {len(data['ifta_trips'])} for division")

            if "speeding" in enabled:
                print("    Fetching speeding events...")
                all_speeding = self.motive.get_speeding_events(start_dt, end_dt)
                data["speeding_events"] = self._filter_events_to_vehicles(all_speeding, div_vehicle_nums)
                print(f"    Speeding: {len(all_speeding)} total -> {len(data['speeding_events'])} for division")

            if "camera_events" in enabled:
                print("    Fetching camera events...")
                all_camera = self.motive.get_camera_events(start_dt, end_dt)
                data["camera_events"] = self._filter_events_to_vehicles(all_camera, div_vehicle_nums)
                print(f"    Camera: {len(all_camera)} total -> {len(data['camera_events'])} for division")

        # -- Man-hours --
        if enabled.intersection({"incidents", "executive_summary"}):
            print("    Loading man-hours...")
            month_str = f"{year}-{month:02d}"
            # Auto-detect file format
            if self.man_hours.has_detail_sheet():
                # 2026 format: single Detail sheet with all companies
                div_key = None
                from .config import DIVISIONS
                for k, v in DIVISIONS.items():
                    if v is self.config:
                        div_key = k
                        break
                data["man_hours"] = self.man_hours.parse_2026(month_str, division_key=div_key)
            else:
                # 2025 format: separate sheets per company
                sheet = self.config.get("man_hours_sheet", "")
                div_filter = self.config.get("man_hours_division")
                if sheet:
                    data["man_hours"] = self.man_hours.parse(sheet, div_filter)
                else:
                    data["man_hours"] = ManHoursParser._empty_result()

        print(f"    Data collection complete.")
        return data

    def collect_quarterly(self, year, quarter):
        """Collect data for a quarterly report (3 months + trends)."""
        months = {1: [1, 2, 3], 2: [4, 5, 6], 3: [7, 8, 9], 4: [10, 11, 12]}[quarter]

        monthly_data = []
        for m in months:
            monthly_data.append(self.collect_monthly(year, m))

        return {
            "division": self.config,
            "year": year,
            "quarter": quarter,
            "period": "quarterly",
            "months": monthly_data,
        }

    def collect_yearly(self, year):
        """Collect data for a yearly report (12 months + full year trends)."""
        monthly_data = []
        for m in range(1, 13):
            monthly_data.append(self.collect_monthly(year, m))

        return {
            "division": self.config,
            "year": year,
            "period": "yearly",
            "months": monthly_data,
        }

    # --- Internal helpers ---

    def _fetch_form(self, form_id, start, end):
        """Fetch form responses and filter to this division."""
        if not form_id:
            return ([], [])
        headers, rows = self.kpa.get_form_responses(form_id, start, end)
        return (headers, self._filter_rows(rows))

    def _filter_rows(self, rows):
        """Filter KPA CSV rows to this division using company + service line.

        KPA form responses use hash-based field names. The company and service line
        fields vary by form, so we check both human-readable names AND known hash IDs.

        Known field hashes for company/service line across forms:
        - Observation 151085: t5187momol3em85v (company), 64c7upqkyt79zhh1 (service line)
        - Incident 151622: lsx3msa0w9n9edb4 (company), sha7vur5q2l6d6gq (service line)
        - Field Assessment 381707: (varies)
        - Transcend obs 484193: t5187momol3em85v (company)
        """
        company_filter = self.config.get("kpa_company_filter", "").lower()
        service_line = self.config.get("kpa_service_line", "").lower()
        alts = [s.lower() for s in self.config.get("kpa_service_line_alt", [])]

        # If no filter configured, return all rows
        if not company_filter and not service_line:
            return rows

        # Hash field IDs that commonly contain company name
        COMPANY_KEYS = (
            "company", "Company", "company_name", "Company Name",
            "t5187momol3em85v",  # Observation card company
            "lsx3msa0w9n9edb4",  # Incident reporting company
            "ge09m6h1ne6po6x9",  # BTI inspection company
            "6zyx6l5f244mk0v5",  # JSA Log company
        )
        # Hash field IDs that commonly contain service line
        SERVICE_LINE_KEYS = (
            "service_line", "Service Line", "service_line_name", "line_of_business",
            "64c7upqkyt79zhh1",  # Observation card service line
            "sha7vur5q2l6d6gq",  # Incident reporting service line
            "77ykc2bzrss3qvxy",  # JSA Log / Transcend service line
            "hxy6pwclvjke1sln",  # Vehicle inspection service line
        )

        filtered = []
        for row in rows:
            # Find company value
            company = ""
            for key in COMPANY_KEYS:
                val = row.get(key, "")
                if val:
                    company = val.strip().lower()
                    break

            # Find service line value
            svc = ""
            for key in SERVICE_LINE_KEYS:
                val = row.get(key, "")
                if val:
                    svc = val.strip().lower()
                    break

            # Match: company contains filter AND (service line matches OR is in alts)
            if company_filter and company_filter not in company:
                continue
            if service_line and svc:
                svc_match = (svc == service_line or svc in alts
                             or svc.startswith(service_line)
                             or any(svc.startswith(a) for a in alts))
                if not svc_match:
                    continue

            filtered.append(row)

        return filtered

    def _fetch_followups(self):
        """Fetch follow-ups filtered to this division.

        Followups have form_id but no company/service_line field.
        We filter by form_id: keep followups from shared HSE forms
        (which all divisions use) AND division-specific forms.

        Then we further filter shared-form followups using m_observer_id
        against division user IDs when available.
        """
        headers, rows = self.kpa.get_followups()

        # Build set of division-specific form IDs
        div_form_ids = set()
        for key, val in self.config.get("form_ids", {}).items():
            if isinstance(val, list):
                div_form_ids.update(val)
            elif isinstance(val, int):
                div_form_ids.add(val)

        shared_form_ids = set(SHARED_FORMS.values())

        # Get division user IDs for filtering shared-form followups
        div_user_ids = self._get_division_user_ids()

        filtered = []
        for row in rows:
            fid = row.get("form_id")
            if fid in div_form_ids:
                # Division-specific form -- always include
                filtered.append(row)
            elif fid in shared_form_ids:
                # Shared form -- filter by user if we have user IDs
                if div_user_ids:
                    observer = row.get("m_observer_id", "")
                    assigner = row.get("m_assigner_id", "")
                    if observer in div_user_ids or assigner in div_user_ids:
                        filtered.append(row)
                else:
                    # No user filtering possible -- include all shared
                    filtered.append(row)
            # Skip followups from forms not in our config at all

        # Enrich with human-readable names
        user_lookup = self._get_user_lookup()
        for row in filtered:
            assignee_id = row.get("m_assignee_id", "")
            row["assigned_to"] = user_lookup.get(assignee_id, "")
            observer_id = row.get("m_observer_id", "")
            row["observer_name"] = user_lookup.get(observer_id, "")

        return (headers, filtered)

    def _fetch_training_status(self, end_date=None):
        """Fetch training compliance status filtered to this division.

        Training-employee-status returns JSON with m_user_id, status, and
        percent_complete. We filter by matching m_user_id against the
        division's user IDs (from lineOfBusiness_id mapping).

        If end_date is provided (YYYY-MM-DD), training programs created after
        that date are excluded from incomplete counts, AND trainings completed
        after that date are moved from complete back to incomplete. This ensures
        compliance is calculated as of the report end date.
        """
        headers, rows = self.kpa.get_training_employee_status()

        # Filter to division users
        div_user_ids = self._get_division_user_ids()
        if div_user_ids:
            rows = [r for r in rows if r.get("m_user_id") in div_user_ids]

        # Enrich with human-readable names
        user_lookup = self._get_user_lookup()
        training_lookup = self._get_training_lookup()
        training_created = getattr(self, "_training_created", {})

        # Calculate cutoff: exclude trainings created after report end date
        cutoff_ms = None
        cutoff_date_int = None
        post_period_completions = set()  # (user_id, training_id) pairs
        if end_date:
            try:
                dt = datetime.strptime(end_date, "%Y-%m-%d").replace(
                    hour=23, minute=59, second=59, tzinfo=timezone.utc
                )
                cutoff_ms = int(dt.timestamp() * 1000)
                cutoff_date_int = int(end_date.replace("-", ""))
            except ValueError:
                pass

            # Fetch completions after end_date to exclude from "complete" list
            if cutoff_date_int:
                print(f"    Checking for post-period training completions (after {end_date})...")
                _, comp_rows = self.kpa.get_completed_trainings()
                for cr in comp_rows:
                    dn = cr.get("date_number", 0) or 0
                    if dn > cutoff_date_int:
                        uid = cr.get("m_user_id", "")
                        tid = cr.get("training_id")
                        if uid and tid is not None:
                            post_period_completions.add((uid, tid))
                if post_period_completions:
                    print(f"    Found {len(post_period_completions)} post-period completions to exclude")

        # Normalize status field and resolve names for downstream renderers
        for row in rows:
            # Add employee name
            uid = row.get("m_user_id", "")
            row["employee_name"] = user_lookup.get(uid, "")

            # Filter and resolve training program names (IDs are ints like 36472)
            incomplete_ids = row.get("incomplete_training_ids", []) or []
            complete_ids = row.get("complete_training_ids", []) or []

            # Exclude trainings created after report period
            if cutoff_ms:
                incomplete_ids = [
                    tid for tid in incomplete_ids
                    if training_created.get(tid, 0) <= cutoff_ms
                ]
                complete_ids = [
                    tid for tid in complete_ids
                    if training_created.get(tid, 0) <= cutoff_ms
                ]

            # Move post-period completions from complete back to incomplete
            if post_period_completions:
                moved = [tid for tid in complete_ids if (uid, tid) in post_period_completions]
                if moved:
                    complete_ids = [tid for tid in complete_ids if tid not in moved]
                    incomplete_ids = incomplete_ids + moved

            row["incomplete_training_names"] = [
                training_lookup.get(tid, f"Program #{tid}") for tid in incomplete_ids
            ]
            row["complete_training_names"] = [
                training_lookup.get(tid, f"Program #{tid}") for tid in complete_ids
            ]

            # Recalculate percent_complete based on filtered lists
            total_trainings = len(incomplete_ids) + len(complete_ids)
            if total_trainings > 0:
                row["percent_complete"] = round(len(complete_ids) / total_trainings * 100)
            else:
                row["percent_complete"] = 100  # No trainings assigned = compliant

            # Set status based on recalculated percent
            pct = row["percent_complete"]
            if pct >= 100:
                row["status"] = "Complete"
            elif row.get("status") == "overdue":
                row["status"] = "Overdue"
            elif pct > 0:
                row["status"] = "In Progress"

        return (headers, rows)

    def _fetch_completed_trainings(self, start, end):
        """Fetch completed trainings in date range, filtered to division users."""
        headers, rows = self.kpa.get_completed_trainings(start, end)
        div_user_ids = self._get_division_user_ids()
        if div_user_ids:
            rows = [r for r in rows if r.get("m_user_id") in div_user_ids]
        return (headers, rows)

    @staticmethod
    def _filter_trips_to_vehicles(trips, vehicle_nums):
        """Filter IFTA trips to only include trips from division vehicles."""
        if not vehicle_nums:
            return trips
        filtered = []
        for trip in trips:
            t = trip.get("ifta_trip_report", trip)
            vehicle = t.get("vehicle", {})
            if isinstance(vehicle, dict):
                vnum = str(vehicle.get("number", ""))
            else:
                vnum = str(vehicle)
            if vnum in vehicle_nums:
                filtered.append(trip)
        return filtered

    @staticmethod
    def _filter_events_to_vehicles(events, vehicle_nums):
        """Filter speeding/camera events to only include division vehicles."""
        if not vehicle_nums:
            return events
        filtered = []
        for evt in events:
            # Speeding events have vehicle.number
            vehicle = evt.get("vehicle", {})
            if isinstance(vehicle, dict):
                vnum = str(vehicle.get("number", ""))
            else:
                vnum = str(vehicle)
            if vnum in vehicle_nums:
                filtered.append(evt)
        return filtered

    def _fetch_assessments(self, start, end):
        """Fetch division-specific safety assessments."""
        form_ids = []

        # Field assessment
        fa = self.config.get("form_ids", {}).get("field_assessment")
        if fa:
            form_ids.append(fa) if not isinstance(fa, list) else form_ids.extend(fa)

        # Management audit
        ma = self.config.get("form_ids", {}).get("management_audit")
        if ma:
            form_ids.append(ma) if not isinstance(ma, list) else form_ids.extend(ma)

        all_rows = []
        all_headers = []
        for fid in form_ids:
            headers, rows = self.kpa.get_form_responses(fid, start, end)
            if not all_headers:
                all_headers = headers
            all_rows.extend(rows)

        return (all_headers, all_rows)
